from __future__ import annotations

import argparse
import io
import json
import os
import re
import unicodedata
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Iterable

from google import genai
from google.genai import types
from openpyxl import load_workbook
from pypdf import PdfReader
import pypdfium2 as pdfium

DEFAULT_MODEL = os.getenv("GEMINI_MODEL", "gemini-3.5-flash")

BASE_DIR = Path(__file__).resolve().parents[1]
REFERENCIAS_FNDE_DIR = BASE_DIR / "data" / "referencias_fnde"
ARQUIVO_REGRAS = REFERENCIAS_FNDE_DIR / "regras_programas.json"
ARQUIVO_ALIASES = REFERENCIAS_FNDE_DIR / "aliases_programas.json"
ARQUIVO_EXCLUSOES = REFERENCIAS_FNDE_DIR / "palavras_excluir.json"
ARQUIVO_CASOS_TESTE = REFERENCIAS_FNDE_DIR / "casos_teste_fnde.json"

COLUNAS_FNDE = {
    "PNAE": 20,   # T
    "PNATE": 21,  # U
    "PDDE": 22,   # V
    "QSE": 23,    # W
}


@dataclass
class OcorrenciaFNDE:
    programa: str
    titulo: str
    valor: float
    pagina: int | None = None
    justificativa: str = ""


@dataclass
class ResultadoFNDE:
    codigo_ibge: str
    municipio: str
    uf: str
    pnae: float = 0.0
    pnate: float = 0.0
    pdde: float = 0.0
    qse: float = 0.0
    ocorrencias: list[OcorrenciaFNDE] = field(default_factory=list)
    avisos: list[str] = field(default_factory=list)

    def valores_planilha(self) -> dict[str, float]:
        return {
            "PNAE": round(float(self.pnae or 0.0), 2),
            "PNATE": round(float(self.pnate or 0.0), 2),
            "PDDE": round(float(self.pdde or 0.0), 2),
            "QSE": round(float(self.qse or 0.0), 2),
        }

    def como_dict(self) -> dict[str, Any]:
        return asdict(self)


CONFIG_AVISOS: list[str] = []


def _carregar_json_configuracao(
    caminho: Path,
    padrao: dict[str, Any],
) -> dict[str, Any]:
    if not caminho.exists():
        CONFIG_AVISOS.append(
            f"Configuração opcional não encontrada: {caminho.as_posix()}. "
            "As regras internas de segurança serão utilizadas."
        )
        return padrao

    try:
        conteudo = json.loads(caminho.read_text(encoding="utf-8"))
    except (OSError, UnicodeError, json.JSONDecodeError) as erro:
        CONFIG_AVISOS.append(
            f"Não foi possível carregar {caminho.name}: {erro}. "
            "As regras internas de segurança serão utilizadas."
        )
        return padrao

    if not isinstance(conteudo, dict):
        CONFIG_AVISOS.append(
            f"O arquivo {caminho.name} não contém um objeto JSON válido. "
            "As regras internas de segurança serão utilizadas."
        )
        return padrao

    return conteudo


_ALIASES_PADRAO = {
    "aliases": {
        "PNAE": [
            "PROGRAMA NACIONAL DE ALIMENTAÇÃO ESCOLAR",
            "PROG.NACIONAL DE ALIMENTAÇÃO ESCOLAR",
            "NACIONAL DE ALIMENTAÇÃO ESCOLAR",
            "PNAE",
        ],
        "PNATE": [
            "PROGRAMA NACIONAL DE APOIO AO TRANSPORTE DO ESCOLAR",
            "PROGRAMA NACIONAL DE APOIO AO TRANSP DO ESCOLAR",
            "NACIONAL DE APOIO AO TRANSPORTE DO ESCOLAR",
            "NACIONAL DE APOIO AO TRANSP DO ESCOLAR",
            "PNATE",
        ],
        "PDDE": [
            "PROGRAMA DINHEIRO DIRETO NA ESCOLA",
            "PDDE",
        ],
        "QSE": [
            "QUOTA ESTADUAL/MUNICIPAL",
            "COTA ESTADUAL/MUNICIPAL",
            "SALÁRIO-EDUCAÇÃO",
            "QSE",
            "QESE",
        ],
    }
}

_EXCLUSOES_PADRAO = {
    "excluir_do_pdde": [
        "ANTIGO PDDE ESTRUTURA",
        "ÁGUA E ESGOTAMENTO SANITÁRIO",
        "ESCOLA DO CAMPO",
        "ESCOLA ACESSÍVEL",
        "PDE ESCOLA",
        "ENSINO MÉDIO INOVADOR",
        "MAIS CULTURA",
        "ESCOLA DE FRONTEIRA",
        "ATLETA NA ESCOLA",
        "ESCOLA SUSTENTÁVEL",
    ],
    "ignorar_programas": [
        "PNLD",
        "PROGRAMA NACIONAL DO LIVRO DIDÁTICO",
    ],
}

_REGRAS_PADRAO = {
    "programas": {
        "PNAE": {"titulo_canonico": "PROGRAMA NACIONAL DE ALIMENTAÇÃO ESCOLAR"},
        "PNATE": {"titulo_canonico": "PROGRAMA NACIONAL DE APOIO AO TRANSPORTE DO ESCOLAR"},
        "PDDE": {"titulo_canonico": "PROGRAMA DINHEIRO DIRETO NA ESCOLA"},
        "QSE": {"titulo_canonico": "QUOTA ESTADUAL/MUNICIPAL"},
    }
}

REGRAS_CONFIG: dict[str, Any] = {}
ALIASES_CONFIG: dict[str, Any] = {}
EXCLUSOES_CONFIG: dict[str, Any] = {}
CASOS_TESTE_CONFIG: dict[str, Any] = {}
ALIASES_NORMALIZADOS: dict[str, tuple[str, ...]] = {}
EXCLUSOES_PDDE_NORMALIZADAS: tuple[str, ...] = ()
PROGRAMAS_IGNORADOS_NORMALIZADOS: tuple[str, ...] = ()


def _lista_textos_configuracao(valor: Any) -> list[str]:
    if not isinstance(valor, list):
        return []
    return [str(item).strip() for item in valor if str(item).strip()]


def _normalizar_lista_configuracao(valores: list[str]) -> tuple[str, ...]:
    return tuple(
        item
        for item in (normalizar_texto(valor) for valor in valores)
        if item
    )


def _aliases_normalizados() -> dict[str, tuple[str, ...]]:
    aliases_brutos = ALIASES_CONFIG.get("aliases", {})
    if not isinstance(aliases_brutos, dict):
        aliases_brutos = {}

    resultado: dict[str, tuple[str, ...]] = {}
    aliases_padrao = _ALIASES_PADRAO["aliases"]

    for programa in ("PNAE", "PNATE", "PDDE", "QSE"):
        configurados = _lista_textos_configuracao(aliases_brutos.get(programa))
        padrao = _lista_textos_configuracao(aliases_padrao.get(programa))
        resultado[programa] = _normalizar_lista_configuracao(
            list(dict.fromkeys(configurados + padrao))
        )

    return resultado


def _inicializar_configuracoes() -> None:
    global REGRAS_CONFIG
    global ALIASES_CONFIG
    global EXCLUSOES_CONFIG
    global CASOS_TESTE_CONFIG
    global ALIASES_NORMALIZADOS
    global EXCLUSOES_PDDE_NORMALIZADAS
    global PROGRAMAS_IGNORADOS_NORMALIZADOS

    REGRAS_CONFIG = _carregar_json_configuracao(ARQUIVO_REGRAS, _REGRAS_PADRAO)
    ALIASES_CONFIG = _carregar_json_configuracao(ARQUIVO_ALIASES, _ALIASES_PADRAO)
    EXCLUSOES_CONFIG = _carregar_json_configuracao(ARQUIVO_EXCLUSOES, _EXCLUSOES_PADRAO)
    CASOS_TESTE_CONFIG = _carregar_json_configuracao(
        ARQUIVO_CASOS_TESTE,
        {"casos": []},
    )

    ALIASES_NORMALIZADOS = _aliases_normalizados()
    EXCLUSOES_PDDE_NORMALIZADAS = _normalizar_lista_configuracao(
        _lista_textos_configuracao(EXCLUSOES_CONFIG.get("excluir_do_pdde"))
        + _lista_textos_configuracao(_EXCLUSOES_PADRAO.get("excluir_do_pdde"))
    )
    PROGRAMAS_IGNORADOS_NORMALIZADOS = _normalizar_lista_configuracao(
        _lista_textos_configuracao(EXCLUSOES_CONFIG.get("ignorar_programas"))
        + _lista_textos_configuracao(_EXCLUSOES_PADRAO.get("ignorar_programas"))
    )


def _resumo_regras_para_prompt() -> str:
    linhas: list[str] = []
    for programa in ("PNAE", "PNATE", "PDDE", "QSE"):
        aliases = ALIASES_CONFIG.get("aliases", {}).get(programa, [])
        if isinstance(aliases, list):
            nomes = "; ".join(str(item) for item in aliases if str(item).strip())
            if nomes:
                linhas.append(f"- {programa}: {nomes}")

    exclusoes = EXCLUSOES_CONFIG.get("excluir_do_pdde", [])
    if isinstance(exclusoes, list) and exclusoes:
        linhas.append(
            "- Nunca incorporar ao PDDE: "
            + "; ".join(str(item) for item in exclusoes if str(item).strip())
        )

    return "\n".join(linhas)


def _api_key() -> str:
    chave = (
        os.getenv("GEMINI_API_KEY")
        or os.getenv("GOOGLE_API_KEY")
        or ""
    ).strip()

    if not chave:
        raise RuntimeError(
            "Chave do Gemini não encontrada. Configure GEMINI_API_KEY "
            "ou GOOGLE_API_KEY no Render."
        )

    return chave


def normalizar_texto(valor: Any) -> str:
    texto = str(valor or "").strip()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(
        caractere
        for caractere in texto
        if not unicodedata.combining(caractere)
    )
    texto = texto.upper()
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip()


_inicializar_configuracoes()


def normalizar_codigo_ibge(valor: Any) -> str:
    if valor is None:
        return ""

    if isinstance(valor, float):
        valor = int(valor)

    codigo = re.sub(r"\D", "", str(valor))
    return codigo.zfill(7) if codigo else ""


def converter_valor_brasileiro(valor: Any) -> float | None:
    if valor is None or valor == "":
        return None

    if isinstance(valor, (int, float)):
        return round(float(valor), 2)

    texto = str(valor).strip()
    texto = re.sub(r"[^\d,.\-]", "", texto)

    if not texto:
        return None

    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")

    try:
        return round(float(texto), 2)
    except ValueError:
        return None


def extrair_identificacao_nome_pdf(
    caminho_pdf: str | Path,
) -> tuple[str, str, str]:
    nome = Path(caminho_pdf).stem

    codigo_match = re.search(r"(?<!\d)(\d{7})(?!\d)", nome)
    codigo_ibge = codigo_match.group(1) if codigo_match else ""

    uf_match = re.search(r"[-_/\s]([A-Z]{2})\s*$", nome.upper())
    uf = uf_match.group(1) if uf_match else ""

    municipio = ""
    if codigo_match:
        municipio = nome[codigo_match.end():]
        municipio = re.sub(
            r"[-_/\s]+[A-Za-z]{2}\s*$",
            "",
            municipio,
        )
        municipio = municipio.strip(" _-/")

    return codigo_ibge, municipio, uf


def extrair_texto_pdf(caminho_pdf: str | Path) -> str:
    leitor = PdfReader(str(caminho_pdf))
    paginas: list[str] = []

    for numero, pagina in enumerate(leitor.pages, start=1):
        try:
            texto = pagina.extract_text() or ""
        except Exception:
            texto = ""

        paginas.append(
            f"\n===== PÁGINA {numero} =====\n{texto.strip()}"
        )

    return "\n".join(paginas).strip()


def renderizar_paginas_png(
    caminho_pdf: str | Path,
    escala: float = 1.7,
    limite_paginas: int = 20,
) -> list[tuple[int, bytes]]:
    documento = pdfium.PdfDocument(str(caminho_pdf))
    imagens: list[tuple[int, bytes]] = []

    try:
        total = min(len(documento), limite_paginas)

        for indice in range(total):
            pagina = documento[indice]
            bitmap = pagina.render(scale=escala)
            imagem = bitmap.to_pil()

            buffer = io.BytesIO()
            imagem.save(buffer, format="PNG", optimize=True)
            imagens.append((indice + 1, buffer.getvalue()))

            imagem.close()
            bitmap.close()
            pagina.close()
    finally:
        documento.close()

    return imagens


def _normalizar_lista_aliases(programa: str) -> list[str]:
    aliases_json = (
        CONFIG_FNDE.get("aliases", {})
        .get("aliases", {})
        .get(programa, [])
    )

    aliases_padrao = {
        "PNAE": [
            "PROGRAMA NACIONAL DE ALIMENTACAO ESCOLAR",
            "PROG NACIONAL DE ALIMENTACAO ESCOLAR",
            "NACIONAL DE ALIMENTACAO ESCOLAR",
            "PNAE",
        ],
        "PNATE": [
            "PROGRAMA NACIONAL DE APOIO AO TRANSPORTE DO ESCOLAR",
            "PROGRAMA NACIONAL DE APOIO AO TRANSP DO ESCOLAR",
            "NACIONAL DE APOIO AO TRANSPORTE DO ESCOLAR",
            "NACIONAL DE APOIO AO TRANSP DO ESCOLAR",
            "PNATE",
        ],
        "PDDE": [
            "PROGRAMA DINHEIRO DIRETO NA ESCOLA",
            "PDDE",
        ],
        "QSE": [
            "QUOTA ESTADUAL MUNICIPAL",
            "COTA ESTADUAL MUNICIPAL",
            "SALARIO EDUCACAO",
            "QSE",
            "QESE",
        ],
    }

    combinados = [
        normalizar_texto(item)
        for item in [*aliases_padrao.get(programa, []), *aliases_json]
        if str(item).strip()
    ]
    return sorted(set(combinados), key=len, reverse=True)


def _expressoes_excluidas_pdde() -> list[str]:
    exclusoes_json = (
        CONFIG_FNDE.get("exclusoes", {})
        .get("excluir_do_pdde", [])
    )
    exclusoes_padrao = [
        "ANTIGO PDDE ESTRUTURA",
        "AGUA E ESGOTAMENTO SANITARIO",
        "ESCOLA DO CAMPO",
        "ESCOLA ACESSIVEL",
        "PDE ESCOLA",
        "ENSINO MEDIO INOVADOR",
        "MAIS CULTURA",
        "ESCOLA DE FRONTEIRA",
        "ATLETA NA ESCOLA",
        "ESCOLA SUSTENTAVEL",
    ]
    return sorted(
        {
            normalizar_texto(item)
            for item in [*exclusoes_padrao, *exclusoes_json]
            if str(item).strip()
        },
        key=len,
        reverse=True,
    )


ALIASES_FNDE = {
    programa: _normalizar_lista_aliases(programa)
    for programa in ("PNAE", "PNATE", "PDDE", "QSE")
}
EXCLUSOES_PDDE = _expressoes_excluidas_pdde()


def _classificar_titulo_programa(titulo: str) -> str | None:
    texto = normalizar_texto(titulo)

    if any(expressao in texto for expressao in EXCLUSOES_PDDE):
        return None

    for programa in ("PNAE", "PNATE", "PDDE", "QSE"):
        if any(alias in texto for alias in ALIASES_FNDE[programa]):
            return programa

    return None


def _recalcular_valores_por_ocorrencias(
    ocorrencias: list[OcorrenciaFNDE],
) -> tuple[dict[str, float], list[OcorrenciaFNDE], list[str]]:
    totais = {"PNAE": 0.0, "PNATE": 0.0, "PDDE": 0.0, "QSE": 0.0}
    ocorrencias_validas: list[OcorrenciaFNDE] = []
    avisos: list[str] = []
    vistos: set[tuple[str, str, int | None, float]] = set()

    for ocorrencia in ocorrencias:
        programa = _classificar_titulo_programa(ocorrencia.titulo)
        if programa is None:
            avisos.append(
                "Ocorrência descartada por não pertencer a um bloco oficial: "
                f"{ocorrencia.titulo or 'sem título'}."
            )
            continue

        chave = (
            programa,
            normalizar_texto(ocorrencia.titulo),
            ocorrencia.pagina,
            round(float(ocorrencia.valor), 2),
        )
        if chave in vistos:
            avisos.append(
                f"Ocorrência duplicada descartada: {ocorrencia.titulo}."
            )
            continue

        vistos.add(chave)
        ocorrencia.programa = programa
        ocorrencias_validas.append(ocorrencia)
        totais[programa] += round(float(ocorrencia.valor), 2)

    return (
        {programa: round(valor, 2) for programa, valor in totais.items()},
        ocorrencias_validas,
        avisos,
    )


def _resumo_regras_para_prompt() -> str:
    return f"""
ALIASES CARREGADOS DOS JSONs:
- PNAE: {ALIASES_FNDE['PNAE']}
- PNATE: {ALIASES_FNDE['PNATE']}
- PDDE: {ALIASES_FNDE['PDDE']}
- QSE: {ALIASES_FNDE['QSE']}

EXCLUSÕES OBRIGATÓRIAS DO PDDE:
{EXCLUSOES_PDDE}
""".strip()


def _schema_gemini() -> dict[str, Any]:
    return {
        "type": "object",
        "properties": {
            "codigo_ibge": {"anyOf": [{"type": "string"}, {"type": "null"}]},
            "municipio": {"anyOf": [{"type": "string"}, {"type": "null"}]},
            "uf": {"anyOf": [{"type": "string"}, {"type": "null"}]},
            "pnae": {"anyOf": [{"type": "number"}, {"type": "null"}]},
            "pnate": {"anyOf": [{"type": "number"}, {"type": "null"}]},
            "pdde": {"anyOf": [{"type": "number"}, {"type": "null"}]},
            "qse": {"anyOf": [{"type": "number"}, {"type": "null"}]},
            "ocorrencias": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "programa": {
                            "type": "string",
                            "enum": ["PNAE", "PNATE", "PDDE", "QSE"],
                        },
                        "titulo": {"type": "string"},
                        "valor": {"type": "number"},
                        "pagina": {"anyOf": [{"type": "integer"}, {"type": "null"}]},
                        "justificativa": {"type": "string"},
                    },
                    "required": [
                        "programa", "titulo", "valor", "pagina", "justificativa"
                    ],
                    "additionalProperties": False,
                },
            },
            "avisos": {"type": "array", "items": {"type": "string"}},
        },
        "required": [
            "codigo_ibge", "municipio", "uf", "pnae", "pnate", "pdde",
            "qse", "ocorrencias", "avisos"
        ],
        "additionalProperties": False,
    }


def _prompt_fnde(
    nome_arquivo: str,
    codigo_ibge: str,
    municipio: str,
    uf: str,
    texto_pdf: str,
) -> str:
    return f"""
Você é um extrator especializado nos demonstrativos de Liberações do
FNDE/SIGEFWEB.

ARQUIVO:
- nome: {nome_arquivo}
- código IBGE identificado no nome: {codigo_ibge or 'não identificado'}
- município identificado no nome: {municipio or 'não identificado'}
- UF identificada no nome: {uf or 'não identificada'}

OBJETIVO:
Extrair os valores totais de PNAE, PNATE, PDDE e QSE/QESE e retornar JSON.

REGRAS OBRIGATÓRIAS:
1. Cada programa deve ser extraído apenas do seu próprio bloco.
2. Use o "Valor Total" exibido no título do bloco; as linhas das esferas
   servem somente para conferência e não podem ser somadas novamente.
3. PNAE: aceitar as variações configuradas para o Programa Nacional de
   Alimentação Escolar.
4. PNATE: aceitar as variações configuradas para o Programa Nacional de
   Apoio ao Transporte do Escolar.
5. PDDE: aceitar somente o bloco principal "PROGRAMA DINHEIRO DIRETO NA
   ESCOLA". Nunca incluir ANTIGO PDDE ESTRUTURA nem programas associados.
6. QSE: aceitar QUOTA/COTA ESTADUAL MUNICIPAL, SALÁRIO-EDUCAÇÃO, QSE ou QESE.
7. Se houver vários blocos realmente válidos do mesmo programa, somar o
   Valor Total de cada bloco uma única vez.
8. Para cada bloco incluído, registrar título completo, programa, valor,
   página e justificativa.
9. Se um programa não existir, retornar 0.0.
10. Não inventar valores; dúvidas devem ir para avisos.

{_resumo_regras_para_prompt()}

TEXTO EXTRAÍDO DO PDF:
--- INÍCIO ---
{texto_pdf[:120000]}
--- FIM ---
""".strip()


def _texto_contem_alias(texto: str, alias: str) -> bool:
    if not alias:
        return False

    # Siglas curtas exigem limite de palavra para evitar falsos positivos.
    if len(alias) <= 5 and " " not in alias:
        return bool(re.search(rf"\b{re.escape(alias)}\b", texto))

    return alias in texto


def _classificar_titulo_programa(titulo: str) -> str | None:
    texto = normalizar_texto(titulo)
    if not texto:
        return None

    if any(
        _texto_contem_alias(texto, item)
        for item in PROGRAMAS_IGNORADOS_NORMALIZADOS
    ):
        return None

    # Exclusões prevalecem sobre o alias genérico PDDE.
    if any(
        _texto_contem_alias(texto, item)
        for item in EXCLUSOES_PDDE_NORMALIZADAS
    ):
        return None

    for programa in ("PNAE", "PNATE", "PDDE", "QSE"):
        for alias in ALIASES_NORMALIZADOS.get(programa, ()):
            if _texto_contem_alias(texto, alias):
                return programa

    return None


def _recalcular_valores_por_ocorrencias(
    ocorrencias: list[OcorrenciaFNDE],
) -> tuple[dict[str, float], list[str]]:
    totais = {"PNAE": 0.0, "PNATE": 0.0, "PDDE": 0.0, "QSE": 0.0}
    avisos: list[str] = []
    vistos: set[tuple[str, str, int | None, float]] = set()

    for ocorrencia in ocorrencias:
        programa_correto = _classificar_titulo_programa(ocorrencia.titulo)

        if programa_correto is None:
            avisos.append(
                "Ocorrência descartada por não pertencer a um bloco oficial: "
                f"{ocorrencia.titulo or 'sem título'}."
            )
            continue

        if ocorrencia.programa != programa_correto:
            avisos.append(
                "Classificação corrigida pelo título do bloco: "
                f"{ocorrencia.programa} -> {programa_correto} "
                f"({ocorrencia.titulo})."
            )

        chave = (
            programa_correto,
            normalizar_texto(ocorrencia.titulo),
            ocorrencia.pagina,
            round(float(ocorrencia.valor), 2),
        )
        if chave in vistos:
            avisos.append(
                "Ocorrência duplicada descartada: "
                f"{ocorrencia.titulo} - {ocorrencia.valor:.2f}."
            )
            continue

        vistos.add(chave)
        ocorrencia.programa = programa_correto
        totais[programa_correto] += round(float(ocorrencia.valor), 2)

    return (
        {programa: round(valor, 2) for programa, valor in totais.items()},
        avisos,
    )


def _normalizar_resposta_gemini(
    bruto: dict[str, Any],
    identificacao_nome: tuple[str, str, str],
) -> ResultadoFNDE:
    codigo_nome, municipio_nome, uf_nome = identificacao_nome

    ocorrencias: list[OcorrenciaFNDE] = []
    for item in bruto.get("ocorrencias", []) or []:
        valor = converter_valor_brasileiro(item.get("valor"))
        if valor is None:
            continue

        ocorrencias.append(
            OcorrenciaFNDE(
                programa=str(item.get("programa") or "").upper(),
                titulo=str(item.get("titulo") or "").strip(),
                valor=valor,
                pagina=(int(item["pagina"]) if item.get("pagina") is not None else None),
                justificativa=str(item.get("justificativa") or "").strip(),
            )
        )

    valores_gemini = {
        "PNAE": converter_valor_brasileiro(bruto.get("pnae")) or 0.0,
        "PNATE": converter_valor_brasileiro(bruto.get("pnate")) or 0.0,
        "PDDE": converter_valor_brasileiro(bruto.get("pdde")) or 0.0,
        "QSE": converter_valor_brasileiro(bruto.get("qse")) or 0.0,
    }

    valores_validados, avisos_validacao = _recalcular_valores_por_ocorrencias(
        ocorrencias
    )

    codigo_resposta = normalizar_codigo_ibge(bruto.get("codigo_ibge"))
    codigo = codigo_resposta or codigo_nome
    municipio = str(bruto.get("municipio") or municipio_nome or "").strip()
    uf = str(bruto.get("uf") or uf_nome or "").strip().upper()

    avisos = [
        str(aviso).strip()
        for aviso in (bruto.get("avisos") or [])
        if str(aviso).strip()
    ]
    avisos.extend(AVISOS_CONFIG_FNDE)
    avisos.extend(avisos_validacao)

    for programa in ("PNAE", "PNATE", "PDDE", "QSE"):
        informado = round(float(valores_gemini[programa]), 2)
        validado = round(float(valores_validados[programa]), 2)
        if informado != validado:
            avisos.append(
                f"{programa}: valor informado pelo Gemini ({informado:.2f}) "
                f"substituído pelo valor validado dos blocos ({validado:.2f})."
            )
    avisos.extend(CONFIG_AVISOS)
    avisos.extend(avisos_validacao)

    for programa in ("PNAE", "PNATE", "PDDE", "QSE"):
        valor_gemini = round(float(valores_gemini[programa]), 2)
        valor_validado = round(float(valores_validados[programa]), 2)
        if valor_gemini != valor_validado:
            avisos.append(
                f"{programa}: total informado pelo Gemini ({valor_gemini:.2f}) "
                f"foi substituído pelo total validado dos blocos "
                f"({valor_validado:.2f})."
            )

    if codigo_nome and codigo_resposta and codigo_nome != codigo_resposta:
        avisos.append(
            "O código IBGE informado pelo Gemini divergiu do nome do arquivo; "
            f"foi mantido o código do nome: {codigo_nome}."
        )
        codigo = codigo_nome

    return ResultadoFNDE(
        codigo_ibge=codigo,
        municipio=municipio,
        uf=uf,
        pnae=valores_validados["PNAE"],
        pnate=valores_validados["PNATE"],
        pdde=valores_validados["PDDE"],
        qse=valores_validados["QSE"],
        ocorrencias=ocorrencias,
        avisos=avisos,
    )


def extrair_com_gemini(
    caminho_pdf: str | Path,
    model: str | None = None,
    enviar_imagens: bool = True,
) -> tuple[ResultadoFNDE, str]:
    caminho = Path(caminho_pdf)
    identificacao = extrair_identificacao_nome_pdf(caminho)
    texto_pdf = extrair_texto_pdf(caminho)

    prompt = _prompt_fnde(
        nome_arquivo=caminho.name,
        codigo_ibge=identificacao[0],
        municipio=identificacao[1],
        uf=identificacao[2],
        texto_pdf=texto_pdf,
    )

    conteudo: list[Any] = [prompt]
    texto_fraco = len(normalizar_texto(texto_pdf)) < 800

    if enviar_imagens or texto_fraco:
        for pagina, png in renderizar_paginas_png(caminho):
            conteudo.append(types.Part.from_text(text=f"Imagem da página {pagina}:"))
            conteudo.append(types.Part.from_bytes(data=png, mime_type="image/png"))

    cliente = genai.Client(api_key=_api_key())

    try:
        resposta = cliente.models.generate_content(
            model=model or DEFAULT_MODEL,
            contents=conteudo,
            config=types.GenerateContentConfig(
                response_mime_type="application/json",
                response_json_schema=_schema_gemini(),
            ),
        )

        analisado = getattr(resposta, "parsed", None)
        bruto = analisado if isinstance(analisado, dict) else json.loads(resposta.text or "{}")
        resultado = _normalizar_resposta_gemini(bruto, identificacao)
        return resultado, texto_pdf
    finally:
        cliente.close()


def process(
    caminho_pdf: str | Path,
    model: str | None = None,
    enviar_imagens: bool = True,
) -> tuple[dict[str, float], str, ResultadoFNDE]:
    resultado, texto = extrair_com_gemini(
        caminho_pdf,
        model=model,
        enviar_imagens=enviar_imagens,
    )
    return resultado.valores_planilha(), texto, resultado


def localizar_aba_principal(workbook: Any):
    melhor = workbook.active
    melhor_pontuacao = -1

    for worksheet in workbook.worksheets:
        pontuacao = 0
        for linha in range(1, min(20, worksheet.max_row) + 1):
            for coluna in range(1, worksheet.max_column + 1):
                valor = normalizar_texto(worksheet.cell(linha, coluna).value)
                if "CODIGO IBGE" in valor:
                    pontuacao += 10
                if "ENTE FEDERADO" in valor:
                    pontuacao += 10
                if "PNAE" in valor:
                    pontuacao += 3
                if "PNATE" in valor:
                    pontuacao += 3
                if "PDDE" in valor:
                    pontuacao += 3
                if "QSE" in valor or "QESE" in valor:
                    pontuacao += 3

        if pontuacao > melhor_pontuacao:
            melhor = worksheet
            melhor_pontuacao = pontuacao

    return melhor


def localizar_coluna_ibge(worksheet: Any) -> int:
    for linha in range(1, min(20, worksheet.max_row) + 1):
        for coluna in range(1, worksheet.max_column + 1):
            valor = normalizar_texto(worksheet.cell(linha, coluna).value)
            if "CODIGO IBGE" in valor:
                return coluna

    raise RuntimeError("A coluna 'Código IBGE' não foi encontrada na planilha.")


def localizar_linha_ibge(worksheet: Any, codigo_ibge: str) -> int:
    coluna_ibge = localizar_coluna_ibge(worksheet)
    codigo_procurado = normalizar_codigo_ibge(codigo_ibge)

    for linha in range(1, worksheet.max_row + 1):
        codigo_linha = normalizar_codigo_ibge(worksheet.cell(linha, coluna_ibge).value)
        if codigo_linha == codigo_procurado:
            return linha

    raise RuntimeError(f"Código IBGE {codigo_procurado} não encontrado na planilha.")


def preencher_resultado_na_planilha(
    caminho_planilha: str | Path,
    resultado: ResultadoFNDE,
    caminho_saida: str | Path | None = None,
    nome_aba: str | None = None,
) -> Path:
    """
    Preenche automaticamente T=PNAE, U=PNATE, V=PDDE e W=QSE.
    Se caminho_saida não for informado, atualiza o próprio arquivo.
    """
    entrada = Path(caminho_planilha)
    saida = Path(caminho_saida) if caminho_saida else entrada

    if entrada.resolve() != saida.resolve():
        saida.write_bytes(entrada.read_bytes())

    workbook = load_workbook(saida)

    try:
        if nome_aba:
            if nome_aba not in workbook.sheetnames:
                raise RuntimeError(f"Aba '{nome_aba}' não encontrada.")
            worksheet = workbook[nome_aba]
        else:
            worksheet = localizar_aba_principal(workbook)

        linha = localizar_linha_ibge(worksheet, resultado.codigo_ibge)

        for programa, valor in resultado.valores_planilha().items():
            coluna = COLUNAS_FNDE[programa]
            celula = worksheet.cell(linha, coluna)
            celula.value = valor
            celula.number_format = '#,##0.00'

        workbook.save(saida)
        return saida
    finally:
        workbook.close()


def processar_pdf_e_preencher_planilha(
    caminho_pdf: str | Path,
    caminho_planilha: str | Path,
    caminho_saida: str | Path | None = None,
    nome_aba: str | None = None,
    model: str | None = None,
) -> ResultadoFNDE:
    _, _, resultado = process(caminho_pdf, model=model, enviar_imagens=True)
    preencher_resultado_na_planilha(
        caminho_planilha=caminho_planilha,
        resultado=resultado,
        caminho_saida=caminho_saida,
        nome_aba=nome_aba,
    )
    return resultado


def processar_varios_pdfs(
    arquivos_pdf: Iterable[str | Path],
    caminho_planilha: str | Path,
    caminho_saida: str | Path | None = None,
    nome_aba: str | None = None,
    model: str | None = None,
) -> tuple[Path, list[ResultadoFNDE], list[dict[str, str]]]:
    """Processa vários PDFs e preenche uma única cópia da planilha."""
    entrada = Path(caminho_planilha)
    saida = Path(caminho_saida) if caminho_saida else entrada

    if entrada.resolve() != saida.resolve():
        saida.write_bytes(entrada.read_bytes())

    workbook = load_workbook(saida)
    resultados: list[ResultadoFNDE] = []
    erros: list[dict[str, str]] = []

    try:
        worksheet = workbook[nome_aba] if nome_aba else localizar_aba_principal(workbook)

        for arquivo in arquivos_pdf:
            caminho_pdf = Path(arquivo)
            try:
                _, _, resultado = process(caminho_pdf, model=model, enviar_imagens=True)
                linha = localizar_linha_ibge(worksheet, resultado.codigo_ibge)

                for programa, valor in resultado.valores_planilha().items():
                    coluna = COLUNAS_FNDE[programa]
                    celula = worksheet.cell(linha, coluna)
                    celula.value = valor
                    celula.number_format = '#,##0.00'

                resultados.append(resultado)
            except Exception as erro:
                erros.append({"arquivo": caminho_pdf.name, "erro": str(erro)})

        workbook.save(saida)
        return saida, resultados, erros
    finally:
        workbook.close()


def _formatar_brl(valor: float) -> str:
    texto = f"{valor:,.2f}"
    return texto.replace(",", "X").replace(".", ",").replace("X", ".")


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Extrai PNAE, PNATE, PDDE e QSE de PDF do FNDE com Gemini e "
            "preenche automaticamente as colunas T, U, V e W."
        )
    )
    parser.add_argument("pdf", help="Caminho do PDF do FNDE.")
    parser.add_argument("--planilha", help="Planilha RREO-TCM+FNDE.")
    parser.add_argument("--saida", help="Caminho da planilha de saída.")
    parser.add_argument(
        "--sem-imagens",
        action="store_true",
        help="Não enviar as imagens das páginas ao Gemini.",
    )
    args = parser.parse_args()

    valores, _, resultado = process(
        args.pdf,
        enviar_imagens=not args.sem_imagens,
    )

    print(json.dumps(resultado.como_dict(), ensure_ascii=False, indent=2))
    print()
    print(f"PNAE:  R$ {_formatar_brl(valores['PNAE'])}")
    print(f"PNATE: R$ {_formatar_brl(valores['PNATE'])}")
    print(f"PDDE:  R$ {_formatar_brl(valores['PDDE'])}")
    print(f"QSE:   R$ {_formatar_brl(valores['QSE'])}")

    if args.planilha:
        destino = preencher_resultado_na_planilha(
            caminho_planilha=args.planilha,
            resultado=resultado,
            caminho_saida=args.saida,
        )
        print(f"Planilha salva em: {destino}")


if __name__ == "__main__":
    main()
