"""Mapeamento fixo para preenchimento da nova planilha RREO/FNDE.

Este módulo concentra SOMENTE a regra de destino dos dados extraídos dos PDFs.
Ele não altera a extração RREO, a extração FNDE, os workers ou a interface.

Aba de destino:
    Sequência_PlanilhaCálculo (2)

Regra principal:
    apenas as colunas declaradas em MAPA_CAMPOS_DESTINO podem ser gravadas.
    Todas as demais colunas da planilha são preservadas.
"""

from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from core.validacao import normalizar_texto


ABA_DESTINO = "Sequência_PlanilhaCálculo (2)"
TOTAL_COLUNAS_PLANILHA = 27  # A até AA
FORMATO_NUMERICO = '#,##0.00'


@dataclass(frozen=True)
class CampoDestino:
    """Define um único campo autorizado para gravação."""

    codigo: str
    coluna: int
    letra: str
    descricao: str
    fonte: str  # RREO ou FNDE

    @property
    def chave(self) -> str:
        return self.codigo.strip().upper()


# =============================================================================
# CAMPOS AUTORIZADOS PARA RECEBER DADOS
# =============================================================================
#
# IMPORTANTE:
# - A declaração original informava "COLUNA 15 LETRA V" para 6.2.1.
# - Excel: coluna 15 = O; coluna 22 = V.
# - Como V já é o destino do PNAE, 6.2.1 foi corrigido para O.
#
# Nenhuma coluna fora deste mapa será preenchida por este módulo.
# =============================================================================

CAMPOS_DESTINO: tuple[CampoDestino, ...] = (
    CampoDestino(
        codigo="1.1",
        coluna=16,
        letra="P",
        descricao=(
            "1.1- Receita Resultante do Imposto sobre a Propriedade "
            "Predial e Territorial Urbana - IPTU"
        ),
        fonte="RREO",
    ),
    CampoDestino(
        codigo="1.2",
        coluna=18,
        letra="R",
        descricao=(
            "1.2- Receita Resultante do Imposto sobre Transmissão "
            "Inter Vivos - ITBI"
        ),
        fonte="RREO",
    ),
    CampoDestino(
        codigo="1.3",
        coluna=19,
        letra="S",
        descricao=(
            "1.3- Receita Resultante do Imposto sobre Serviços de "
            "Qualquer Natureza - ISS"
        ),
        fonte="RREO",
    ),
    CampoDestino(
        codigo="1.4",
        coluna=17,
        letra="Q",
        descricao=(
            "1.4- Receita Resultante do Imposto de Renda Retido na Fonte - IRRF"
        ),
        fonte="RREO",
    ),
    CampoDestino(
        codigo="2.1.1",
        coluna=5,
        letra="E",
        descricao="2.1.1- Parcela referente à CF, art. 159, I, alínea b",
        fonte="RREO",
    ),
    CampoDestino(
        codigo="2.1.2",
        coluna=6,
        letra="F",
        descricao="2.1.2- Parcela referente à CF, art. 159, I, alíneas d e e",
        fonte="RREO",
    ),
    CampoDestino(
        codigo="2.3",
        coluna=8,
        letra="H",
        descricao="2.3- Cota-Parte IPI-Exportação",
        fonte="RREO",
    ),
    CampoDestino(
        codigo="2.4",
        coluna=10,
        letra="J",
        descricao="2.4- Cota-Parte ITR",
        fonte="RREO",
    ),
    CampoDestino(
        codigo="2.2",
        coluna=11,
        letra="K",
        descricao="2.2- Cota-Parte ICMS",
        fonte="RREO",
    ),
    CampoDestino(
        codigo="2.5",
        coluna=12,
        letra="L",
        descricao="2.5- Cota-Parte IPVA",
        fonte="RREO",
    ),
    CampoDestino(
        codigo="2.6",
        coluna=20,
        letra="T",
        descricao="2.6- Cota-Parte IOF-Ouro (quando existir no PDF)",
        fonte="RREO",
    ),
    CampoDestino(
        codigo="6.1.1",
        coluna=14,
        letra="N",
        descricao="6.1.1 - Principal",
        fonte="RREO",
    ),
    CampoDestino(
        codigo="6.2.1",
        coluna=15,
        letra="O",
        descricao="6.2.1 - Principal",
        fonte="RREO",
    ),
    CampoDestino(
        codigo="PNAE",
        coluna=22,
        letra="V",
        descricao="Programas Universais - PNAE (Base VAAT)",
        fonte="FNDE",
    ),
    CampoDestino(
        codigo="PNATE",
        coluna=23,
        letra="W",
        descricao="Programas Universais - PNATE (Base VAAT)",
        fonte="FNDE",
    ),
    CampoDestino(
        codigo="PDDE",
        coluna=24,
        letra="X",
        descricao="Programas Universais - PDDE (Base VAAT)",
        fonte="FNDE",
    ),
    CampoDestino(
        codigo="QSE",
        coluna=25,
        letra="Y",
        descricao="Programas Universais - QSE (Base VAAT)",
        fonte="FNDE",
    ),
)

MAPA_CAMPOS_DESTINO: dict[str, CampoDestino] = {
    campo.chave: campo for campo in CAMPOS_DESTINO
}

COLUNAS_AUTORIZADAS: frozenset[int] = frozenset(
    campo.coluna for campo in CAMPOS_DESTINO
)

LETRAS_AUTORIZADAS: frozenset[str] = frozenset(
    campo.letra for campo in CAMPOS_DESTINO
)

CAMPOS_RREO_AUTORIZADOS: frozenset[str] = frozenset(
    campo.chave for campo in CAMPOS_DESTINO if campo.fonte == "RREO"
)

CAMPOS_FNDE_AUTORIZADOS: frozenset[str] = frozenset(
    campo.chave for campo in CAMPOS_DESTINO if campo.fonte == "FNDE"
)


# =============================================================================
# VALIDAÇÕES DO MAPEAMENTO
# =============================================================================

def validar_mapeamento() -> None:
    """Valida a configuração antes de qualquer gravação.

    Impede, por exemplo:
    - coluna fora de A:AA;
    - número de coluna incompatível com a letra;
    - dois campos tentando usar a mesma coluna;
    - código duplicado;
    - fonte diferente de RREO/FNDE.
    """

    codigos: set[str] = set()
    colunas: set[int] = set()

    for campo in CAMPOS_DESTINO:
        if not 1 <= campo.coluna <= TOTAL_COLUNAS_PLANILHA:
            raise ValueError(
                f"Campo {campo.codigo}: coluna {campo.coluna} fora do intervalo "
                f"1-{TOTAL_COLUNAS_PLANILHA}."
            )

        letra_real = get_column_letter(campo.coluna)
        if letra_real != campo.letra.upper():
            raise ValueError(
                f"Campo {campo.codigo}: coluna {campo.coluna} corresponde a "
                f"{letra_real}, mas o mapa informa {campo.letra}."
            )

        if campo.chave in codigos:
            raise ValueError(f"Código duplicado no mapa: {campo.codigo}")
        codigos.add(campo.chave)

        if campo.coluna in colunas:
            raise ValueError(
                f"Coluna {campo.coluna} ({campo.letra}) foi atribuída a mais "
                "de um campo."
            )
        colunas.add(campo.coluna)

        if campo.fonte not in {"RREO", "FNDE"}:
            raise ValueError(
                f"Campo {campo.codigo}: fonte inválida {campo.fonte!r}."
            )


def validar_worksheet(worksheet: Worksheet) -> None:
    """Confirma que a gravação está ocorrendo na aba correta."""

    if worksheet.title != ABA_DESTINO:
        raise ValueError(
            f"Aba incorreta: {worksheet.title!r}. "
            f"A aba autorizada é {ABA_DESTINO!r}."
        )

    # Não exigimos max_column == 27, pois o openpyxl pode considerar colunas
    # adicionais formatadas. O que importa é que as colunas do mapa estejam
    # dentro do limite oficial A:AA.
    if max(COLUNAS_AUTORIZADAS) > TOTAL_COLUNAS_PLANILHA:
        raise ValueError("O mapa contém coluna fora do intervalo A:AA.")


def obter_aba_destino(workbook: Any) -> Worksheet:
    """Obtém exclusivamente a aba configurada para receber os dados."""

    if ABA_DESTINO not in workbook.sheetnames:
        disponiveis = ", ".join(workbook.sheetnames)
        raise KeyError(
            f"A aba {ABA_DESTINO!r} não existe na planilha. "
            f"Abas disponíveis: {disponiveis}"
        )

    worksheet = workbook[ABA_DESTINO]
    validar_worksheet(worksheet)
    return worksheet



# =============================================================================
# ÍNDICE OFICIAL DE MUNICÍPIOS NA NOVA PLANILHA
# =============================================================================

COLUNA_CODIGO_IBGE = 3  # C
COLUNA_ENTE_FEDERADO = 4  # D
PRIMEIRA_LINHA_DADOS = 3


def normalizar_codigo_ibge(valor: Any) -> str:
    """Normaliza Código IBGE vindo do Excel sem perder zeros significativos."""
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    texto = str(valor).strip()
    if texto.endswith(".0") and texto[:-2].isdigit():
        texto = texto[:-2]
    return "".join(ch for ch in texto if ch.isdigit())


def validar_estrutura_planilha(worksheet: Worksheet) -> None:
    """Valida aba, cabeçalhos fixos e as 27 colunas da matriz oficial."""
    validar_worksheet(worksheet)
    if worksheet.max_column < TOTAL_COLUNAS_PLANILHA:
        raise ValueError(
            f"A aba {ABA_DESTINO!r} possui apenas {worksheet.max_column} colunas; "
            f"eram esperadas ao menos {TOTAL_COLUNAS_PLANILHA} (A:AA)."
        )
    cabec_ibge = str(worksheet.cell(1, COLUNA_CODIGO_IBGE).value or "").strip().lower()
    cabec_ente = str(worksheet.cell(1, COLUNA_ENTE_FEDERADO).value or "").strip().lower()
    if "ibge" not in cabec_ibge:
        raise ValueError("A coluna C não contém o cabeçalho de Código IBGE esperado.")
    if not any(token in cabec_ente for token in ("ente", "munic")):
        raise ValueError("A coluna D não contém o cabeçalho de Ente Federado/Município esperado.")


def construir_indice_linhas_ibge(worksheet: Worksheet) -> dict[str, int]:
    """Cria índice Código IBGE -> linha usando exclusivamente a coluna C."""
    validar_estrutura_planilha(worksheet)
    indice: dict[str, int] = {}
    duplicados: list[str] = []
    for row in range(PRIMEIRA_LINHA_DADOS, worksheet.max_row + 1):
        codigo = normalizar_codigo_ibge(worksheet.cell(row, COLUNA_CODIGO_IBGE).value)
        if len(codigo) != 7:
            continue
        if codigo in indice:
            duplicados.append(codigo)
            continue
        indice[codigo] = row
    if duplicados:
        raise ValueError(
            "Códigos IBGE duplicados na planilha: " + ", ".join(sorted(set(duplicados))[:20])
        )
    return indice


def localizar_linha_por_codigo_ibge(worksheet: Worksheet, codigo_ibge: Any) -> int:
    """Localiza a linha oficial do município pela coluna C."""
    codigo = normalizar_codigo_ibge(codigo_ibge)
    if len(codigo) != 7:
        raise ValueError(f"Código IBGE municipal inválido: {codigo_ibge!r}")
    indice = construir_indice_linhas_ibge(worksheet)
    if codigo not in indice:
        raise KeyError(f"Código IBGE {codigo} não encontrado na aba {ABA_DESTINO!r}.")
    return indice[codigo]


def carregar_municipios_da_planilha(
    worksheet: Worksheet,
    uf: str,
    prefixo_uf: str,
) -> list[dict[str, Any]]:
    """Lista municípios usando somente C (IBGE) e D (Ente Federado)."""
    validar_estrutura_planilha(worksheet)
    uf = str(uf or "").upper().strip()
    prefixo_uf = str(prefixo_uf or "").strip()
    municipios: list[dict[str, Any]] = []
    for row in range(PRIMEIRA_LINHA_DADOS, worksheet.max_row + 1):
        codigo = normalizar_codigo_ibge(worksheet.cell(row, COLUNA_CODIGO_IBGE).value)
        if len(codigo) != 7 or not codigo.startswith(prefixo_uf):
            continue
        ente = str(worksheet.cell(row, COLUNA_ENTE_FEDERADO).value or "").strip()
        if not ente:
            continue
        partes = ente.rsplit("/", 1)
        if len(partes) != 2 or partes[1].strip().upper() != uf:
            continue
        nome = partes[0].strip()
        municipios.append({
            "row": row,
            "codigo_ibge": codigo,
            "nome": nome,
            "uf": uf,
            "ente": ente,
            "nome_normalizado": normalizar_texto(nome),
        })
    return municipios

# =============================================================================
# UTILITÁRIOS DE CÉLULA E VALOR
# =============================================================================

def obter_celula_destino(linha_municipio: int, codigo: str) -> str:
    """Retorna a célula exata autorizada para um código, ex.: P5 ou V5601."""

    if linha_municipio < 1:
        raise ValueError("A linha do município deve ser maior ou igual a 1.")

    chave = str(codigo).strip().upper()
    campo = MAPA_CAMPOS_DESTINO.get(chave)
    if campo is None:
        raise KeyError(
            f"Código/campo {codigo!r} não está autorizado para preenchimento."
        )

    return f"{campo.letra}{linha_municipio}"


def _converter_valor_numerico(valor: Any) -> float | int | None:
    """Normaliza valores numéricos sem transformar ausência em zero."""

    if valor is None:
        return None

    if isinstance(valor, bool):
        raise TypeError("Valor booleano não é aceito como valor financeiro.")

    if isinstance(valor, (int, float, Decimal)):
        return float(valor)

    if isinstance(valor, str):
        texto = valor.strip()
        if not texto:
            return None

        # Aceita formato brasileiro: 1.234.567,89
        texto = texto.replace("R$", "").replace(" ", "")
        if "," in texto:
            texto = texto.replace(".", "").replace(",", ".")
        try:
            return float(texto)
        except ValueError as exc:
            raise ValueError(f"Valor financeiro inválido: {valor!r}") from exc

    raise TypeError(
        f"Tipo de valor não suportado: {type(valor).__name__}."
    )


# =============================================================================
# GRAVAÇÃO PRINCIPAL
# =============================================================================

def gravar_dados_mapeados(
    worksheet: Worksheet,
    linha_municipio: int,
    valores: Mapping[str, Any],
    *,
    preservar_formulas: bool = True,
) -> dict[str, Any]:
    """Grava somente campos explicitamente autorizados no mapa.

    Args:
        worksheet: aba já aberta do openpyxl.
        linha_municipio: linha exata do município na nova planilha.
        valores: dicionário com códigos RREO/FNDE e valores extraídos.
        preservar_formulas: se True, não sobrescreve fórmula existente.

    Returns:
        Relatório com quantidade e células alteradas/ignoradas.

    Observações:
        - campos desconhecidos são IGNORADOS;
        - valores None/vazios são IGNORADOS;
        - zero é valor válido e é gravado;
        - nenhuma coluna fora do mapa pode ser alterada por esta função.
    """

    validar_worksheet(worksheet)

    if linha_municipio < 1:
        raise ValueError("A linha do município deve ser maior ou igual a 1.")

    relatorio: dict[str, Any] = {
        "preenchidos": 0,
        "celulas_preenchidas": [],
        "campos_sem_valor": [],
        "campos_nao_mapeados": [],
        "formulas_preservadas": [],
    }

    for codigo_recebido, valor_recebido in valores.items():
        chave = str(codigo_recebido).strip().upper()
        campo = MAPA_CAMPOS_DESTINO.get(chave)

        if campo is None:
            relatorio["campos_nao_mapeados"].append(str(codigo_recebido))
            continue

        valor = _converter_valor_numerico(valor_recebido)
        if valor is None:
            relatorio["campos_sem_valor"].append(campo.codigo)
            continue

        # Segurança adicional: a coluna precisa continuar autorizada.
        if campo.coluna not in COLUNAS_AUTORIZADAS:
            raise RuntimeError(
                f"Tentativa bloqueada de escrever na coluna {campo.coluna}."
            )

        celula = worksheet.cell(
            row=linha_municipio,
            column=campo.coluna,
        )

        if (
            preservar_formulas
            and isinstance(celula.value, str)
            and celula.value.startswith("=")
        ):
            relatorio["formulas_preservadas"].append(celula.coordinate)
            continue

        celula.value = valor
        celula.number_format = FORMATO_NUMERICO

        relatorio["preenchidos"] += 1
        relatorio["celulas_preenchidas"].append(celula.coordinate)

    return relatorio


def preencher_resultados_nova_planilha(
    worksheet: Worksheet,
    linha_municipio: int,
    resultados: Mapping[str, Any],
) -> int:
    """Wrapper simples compatível com o padrão atual do painel.

    Retorna apenas o número de células preenchidas.
    """

    relatorio = gravar_dados_mapeados(
        worksheet,
        linha_municipio,
        resultados,
    )
    return int(relatorio["preenchidos"])


def preencher_rreo_nova_planilha(
    worksheet: Worksheet,
    linha_municipio: int,
    valores_rreo: Mapping[str, Any],
) -> int:
    """Grava somente os códigos RREO autorizados."""

    filtrados = {
        str(codigo).strip().upper(): valor
        for codigo, valor in valores_rreo.items()
        if str(codigo).strip().upper() in CAMPOS_RREO_AUTORIZADOS
    }
    return preencher_resultados_nova_planilha(
        worksheet,
        linha_municipio,
        filtrados,
    )


def preencher_fnde_nova_planilha(
    worksheet: Worksheet,
    linha_municipio: int,
    valores_fnde: Mapping[str, Any],
) -> int:
    """Grava somente PNAE, PNATE, PDDE e QSE nas colunas V:Y."""

    filtrados = {
        str(codigo).strip().upper(): valor
        for codigo, valor in valores_fnde.items()
        if str(codigo).strip().upper() in CAMPOS_FNDE_AUTORIZADOS
    }
    return preencher_resultados_nova_planilha(
        worksheet,
        linha_municipio,
        filtrados,
    )


# =============================================================================
# USO ISOLADO EM UM ARQUIVO XLSX
# =============================================================================

def gravar_em_arquivo(
    caminho_planilha: str | Path,
    linha_municipio: int,
    valores: Mapping[str, Any],
    *,
    preservar_formulas: bool = True,
) -> dict[str, Any]:
    """Abre, grava e salva uma planilha.

    No processamento em lote do Streamlit, prefira abrir o workbook uma única
    vez e usar gravar_dados_mapeados()/preencher_*_nova_planilha().
    """

    caminho = Path(caminho_planilha)
    if not caminho.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {caminho}")

    workbook = load_workbook(caminho, data_only=False)
    try:
        worksheet = obter_aba_destino(workbook)
        relatorio = gravar_dados_mapeados(
            worksheet,
            linha_municipio,
            valores,
            preservar_formulas=preservar_formulas,
        )
        workbook.save(caminho)
        return relatorio
    finally:
        workbook.close()


# Validação automática: se alguém editar o mapa errado, o app falha cedo e
# mostra claramente o erro em vez de preencher uma coluna incorreta.
validar_mapeamento()


__all__ = [
    "ABA_DESTINO",
    "TOTAL_COLUNAS_PLANILHA",
    "CAMPOS_DESTINO",
    "MAPA_CAMPOS_DESTINO",
    "COLUNAS_AUTORIZADAS",
    "CAMPOS_RREO_AUTORIZADOS",
    "CAMPOS_FNDE_AUTORIZADOS",
    "validar_mapeamento",
    "validar_worksheet",
    "obter_aba_destino",
    "obter_celula_destino",
    "gravar_dados_mapeados",
    "preencher_resultados_nova_planilha",
    "preencher_rreo_nova_planilha",
    "preencher_fnde_nova_planilha",
    "gravar_em_arquivo",
]
