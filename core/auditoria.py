"""Fachada de compatibilidade para auditorias separadas por fonte."""
from __future__ import annotations

from datetime import datetime
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from core.auditoria_rreo import append_rreo_log
from core.auditoria_fnde import append_fnde_log

MISSING_SHEET = "MUNICIPIOS_NAO_ENCONTRADOS"
AUDIT_SHEET = "AUDITORIA"
MISSING_HEADERS = [
    "Estado/UF", "Código IBGE", "Município da planilha-base", "Situação",
    "PDF RREO correspondente", "PDF FNDE correspondente", "Observação",
]


def append_log(workbook: Workbook, item: dict[str, Any]) -> None:
    append_rreo_log(workbook, item)


def write_missing(workbook: Workbook, rows: Iterable[dict[str, Any]]) -> None:
    if MISSING_SHEET in workbook.sheetnames:
        del workbook[MISSING_SHEET]
    ws = workbook.create_sheet(MISSING_SHEET)
    ws.append(MISSING_HEADERS)
    for item in rows:
        ws.append([item.get(header, "") for header in MISSING_HEADERS])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    ws.freeze_panes = "A2"


def write_audit(workbook: Workbook, metrics: dict[str, Any]) -> None:
    if AUDIT_SHEET in workbook.sheetnames:
        del workbook[AUDIT_SHEET]
    ws = workbook.create_sheet(AUDIT_SHEET)
    ws.append(["Item", "Valor"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for key, value in metrics.items():
        ws.append([key, value])
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 72
    ws.freeze_panes = "A2"


def timestamp() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
