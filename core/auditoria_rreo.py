from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

RREO_LOG_SHEET = "LOG_RREO"
RREO_LOG_HEADERS = [
    "Data/Hora", "PDF RREO", "Município", "Código IBGE", "UF", "Ano",
    "Estado/Lote", "Linha da planilha", "Município no nome do arquivo",
    "Município no conteúdo", "Divergência de município", "Ação adotada",
    "Origem da identificação", "Confiança da identificação",
    "Campos RREO preenchidos", "Códigos encontrados", "Códigos ausentes",
    "Status", "Método de extração", "Status do upload", "Tentativas de upload",
    "Pasta de destino", "Erro resumido", "Observações",
]


def _sheet(workbook: Workbook):
    if RREO_LOG_SHEET in workbook.sheetnames:
        ws = workbook[RREO_LOG_SHEET]
    else:
        ws = workbook.create_sheet(RREO_LOG_SHEET)
        ws.append(RREO_LOG_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    return ws


def append_rreo_log(workbook: Workbook, item: dict[str, Any]) -> None:
    ws = _sheet(workbook)
    ws.append([item.get(header, "") for header in RREO_LOG_HEADERS])
