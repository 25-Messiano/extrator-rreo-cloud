from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

FNDE_LOG_SHEET = "LOG_FNDE"
FNDE_LOG_HEADERS = [
    "Data/Hora", "PDF FNDE", "Município", "Código IBGE", "UF", "Ano",
    "Estado/Lote", "Linha da planilha", "Campos FNDE preenchidos",
    "Programas encontrados", "Programas ausentes", "Valores extraídos",
    "Método de extração", "Modelo Gemini", "Tentativas Gemini", "Status",
    "Status do upload", "Tentativas de upload", "Pasta de destino",
    "Erro resumido", "Avisos",
]


def _sheet(workbook: Workbook):
    if FNDE_LOG_SHEET in workbook.sheetnames:
        ws = workbook[FNDE_LOG_SHEET]
    else:
        ws = workbook.create_sheet(FNDE_LOG_SHEET)
        ws.append(FNDE_LOG_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    return ws


def append_fnde_log(workbook: Workbook, item: dict[str, Any]) -> None:
    ws = _sheet(workbook)
    ws.append([item.get(header, "") for header in FNDE_LOG_HEADERS])
