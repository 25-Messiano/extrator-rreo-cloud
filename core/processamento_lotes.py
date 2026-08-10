from __future__ import annotations

from concurrent.futures import Future, ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Iterable, TypeVar

from openpyxl import Workbook

T = TypeVar("T")
R = TypeVar("R")
S = TypeVar("S")

CONTROL_SHEET = "__CONTROLE_LOTES"


@dataclass(frozen=True)
class BatchSettings:
    batch_size: int = 10
    rreo_workers: int = 4
    fnde_workers: int = 2
    ui_update_interval: int = 5
    checkpoint_every: int = 10
    resume_enabled: bool = True

    @classmethod
    def from_mapping(cls, values: dict[str, Any]) -> "BatchSettings":
        def positive(name: str, default: int, maximum: int) -> int:
            try:
                value = int(values.get(name, default))
            except (TypeError, ValueError):
                value = default
            return max(1, min(value, maximum))

        return cls(
            batch_size=positive("tamanho_lote", 10, 50),
            rreo_workers=positive("workers_rreo", 4, 8),
            fnde_workers=positive("workers_fnde", 2, 4),
            ui_update_interval=positive("atualizar_tela_cada", 5, 50),
            checkpoint_every=positive("checkpoint_cada", 10, 100),
            resume_enabled=bool(values.get("retomar_processamento", True)),
        )


def chunks(items: list[T], size: int) -> Iterable[list[T]]:
    safe_size = max(1, int(size))
    for start in range(0, len(items), safe_size):
        yield items[start:start + safe_size]


def execute_parallel(
    items: list[T],
    worker: Callable[[T], R],
    max_workers: int,
) -> dict[int, tuple[R | None, Exception | None]]:
    """Executa tarefas em paralelo e preserva a posição original."""
    if not items:
        return {}

    results: dict[int, tuple[R | None, Exception | None]] = {}
    with ThreadPoolExecutor(max_workers=max(1, int(max_workers))) as executor:
        futures: dict[Future[R], int] = {
            executor.submit(worker, item): index
            for index, item in enumerate(items)
        }
        for future in as_completed(futures):
            index = futures[future]
            try:
                results[index] = (future.result(), None)
            except Exception as error:  # worker errors are isolated per item
                results[index] = (None, error)
    return results


def ensure_control_sheet(workbook: Workbook):
    if CONTROL_SHEET in workbook.sheetnames:
        worksheet = workbook[CONTROL_SHEET]
    else:
        worksheet = workbook.create_sheet(CONTROL_SHEET)
        worksheet.append(["chave", "valor"])
    worksheet.sheet_state = "hidden"
    return worksheet


def read_checkpoint_state(workbook: Workbook) -> dict[str, str]:
    if CONTROL_SHEET not in workbook.sheetnames:
        return {}
    worksheet = workbook[CONTROL_SHEET]
    state: dict[str, str] = {}
    for key, value in worksheet.iter_rows(min_row=2, values_only=True):
        if key:
            state[str(key)] = "" if value is None else str(value)
    return state


def write_checkpoint_state(workbook: Workbook, values: dict[str, Any]) -> None:
    worksheet = ensure_control_sheet(workbook)
    worksheet.delete_rows(2, max(worksheet.max_row - 1, 0))
    for key, value in sorted(values.items()):
        worksheet.append([key, "" if value is None else str(value)])
    worksheet.sheet_state = "hidden"


def processed_codes_from_state(state: dict[str, str]) -> set[str]:
    raw = state.get("codigos_processados", "")
    return {item for item in raw.split(",") if item}


def checkpoint_filename(job_id: str, completed: int) -> str:
    safe = "".join(character if character.isalnum() or character in "_-" else "_" for character in job_id)
    return f"CHECKPOINT_{safe}_{completed:06d}.xlsx"


def copy_checkpoint(source: Path, destination: Path) -> Path:
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_bytes(source.read_bytes())
    return destination


def execute_dual_parallel(
    items: list[T],
    rreo_worker: Callable[[T], R] | None,
    fnde_worker: Callable[[T], R] | None,
    rreo_workers: int,
    fnde_workers: int,
) -> tuple[
    dict[int, tuple[R | None, Exception | None]],
    dict[int, tuple[R | None, Exception | None]],
]:
    """Executa dois canais independentes em paralelo, sem escrever no Excel."""
    rreo_results: dict[int, tuple[R | None, Exception | None]] = {}
    fnde_results: dict[int, tuple[R | None, Exception | None]] = {}

    with ThreadPoolExecutor(max_workers=max(1, int(rreo_workers))) as rreo_executor, \
         ThreadPoolExecutor(max_workers=max(1, int(fnde_workers))) as fnde_executor:
        futures: dict[Future[R], tuple[str, int]] = {}
        if rreo_worker is not None:
            futures.update({
                rreo_executor.submit(rreo_worker, item): ("RREO", index)
                for index, item in enumerate(items)
            })
        if fnde_worker is not None:
            futures.update({
                fnde_executor.submit(fnde_worker, item): ("FNDE", index)
                for index, item in enumerate(items)
            })

        for future in as_completed(futures):
            channel, index = futures[future]
            target = rreo_results if channel == "RREO" else fnde_results
            try:
                target[index] = (future.result(), None)
            except Exception as error:
                target[index] = (None, error)

    return rreo_results, fnde_results



def execute_plan_batches(
    tasks: list[dict[str, Any]],
    rreo_worker: Callable[[dict[str, Any]], R] | None,
    fnde_worker: Callable[[dict[str, Any]], S] | None,
    settings: BatchSettings,
):
    """Executa plano já fundido, em lotes, sem escrever Excel nas threads."""
    for number, batch in enumerate(chunks(tasks, settings.batch_size), start=1):
        rreo_results, fnde_results = execute_dual_parallel(
            batch, rreo_worker, fnde_worker,
            settings.rreo_workers, settings.fnde_workers,
        )
        yield number, batch, rreo_results, fnde_results
