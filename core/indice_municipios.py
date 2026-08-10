from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from core.validacao import normalizar_codigo_ibge, normalizar_texto


@dataclass(frozen=True)
class Municipio:
    codigo_ibge: str
    nome: str
    uf: str
    row: int = 0
    ente: str = ""

    @property
    def chave_nome_uf(self) -> tuple[str, str]:
        return normalizar_texto(self.nome), self.uf.upper()

    def as_dict(self) -> dict[str, Any]:
        return {
            "codigo_ibge": self.codigo_ibge,
            "nome": self.nome,
            "uf": self.uf,
            "row": self.row,
            "ente": self.ente,
            "nome_normalizado": normalizar_texto(self.nome),
        }


class IndiceMunicipios:
    def __init__(self, municipios: Iterable[Municipio]):
        self.municipios = list(municipios)
        self.por_ibge = {item.codigo_ibge: item for item in self.municipios if item.codigo_ibge}
        self.por_nome_uf = {item.chave_nome_uf: item for item in self.municipios}
        self.por_uf: dict[str, list[Municipio]] = {}
        for item in self.municipios:
            self.por_uf.setdefault(item.uf.upper(), []).append(item)

    def localizar_nome_uf(self, nome: str, uf: str) -> Municipio | None:
        return self.por_nome_uf.get((normalizar_texto(nome), str(uf).upper()))

    def localizar_ibge(self, codigo: str) -> Municipio | None:
        return self.por_ibge.get(normalizar_codigo_ibge(codigo))

    def do_estado(self, uf: str) -> list[Municipio]:
        return list(self.por_uf.get(str(uf).upper(), []))


def _header_positions(ws) -> tuple[int, int, int]:
    ibge_col = ente_col = header_row = 0
    for row in range(1, min(ws.max_row, 25) + 1):
        for col in range(1, ws.max_column + 1):
            value = normalizar_texto(ws.cell(row, col).value)
            if "CODIGO IBGE" in value:
                ibge_col, header_row = col, row
            elif "ENTE FEDERADO" in value:
                ente_col = col
        if ibge_col and ente_col:
            return ibge_col, ente_col, header_row
    raise RuntimeError("Cabeçalhos Código IBGE e Ente Federado não encontrados.")


def carregar_da_planilha(path: str | Path) -> IndiceMunicipios:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        best = workbook.active
        for ws in workbook.worksheets:
            try:
                _header_positions(ws)
                best = ws
                break
            except RuntimeError:
                continue
        ibge_col, ente_col, header_row = _header_positions(best)
        municipios: list[Municipio] = []
        for row in range(header_row + 1, best.max_row + 1):
            code = normalizar_codigo_ibge(best.cell(row, ibge_col).value)
            ente = str(best.cell(row, ente_col).value or "").strip()
            if len(code) != 7 or not ente:
                continue
            if "/" in ente:
                nome, uf = ente.rsplit("/", 1)
            elif " - " in ente:
                nome, uf = ente.rsplit(" - ", 1)
            else:
                continue
            uf = uf.strip().upper()
            if len(uf) != 2:
                continue
            municipios.append(Municipio(code, nome.strip(), uf, row, ente))
        return IndiceMunicipios(municipios)
    finally:
        workbook.close()


def de_dicionarios(items: Iterable[dict[str, Any]]) -> IndiceMunicipios:
    return IndiceMunicipios(
        Municipio(
            normalizar_codigo_ibge(item.get("codigo_ibge")),
            str(item.get("nome") or "").strip(),
            str(item.get("uf") or "").upper().strip(),
            int(item.get("row") or 0),
            str(item.get("ente") or ""),
        )
        for item in items
        if item.get("nome") and item.get("uf")
    )
