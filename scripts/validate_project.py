from __future__ import annotations

import json
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
PLANILHA = ROOT / "data" / "RREO-TCM+FNDE PLANILHA BASE.xlsx"
ABA = "Sequência_PlanilhaCálculo (2)"

REQUIRED = [
    ROOT / "app.py",
    ROOT / "pages" / "1_Painel.py",
    ROOT / "pages" / "2_Historico.py",
    ROOT / "pages" / "3_Configuracoes.py",
    ROOT / "pages" / "4_Arquivos_Cloud.py",
    ROOT / "pages" / "5_Usuarios.py",
    ROOT / "integrations" / "google_storage.py",
    ROOT / "integrations" / "gemini.py",
    ROOT / "modules" / "rreo.py",
    ROOT / "modules" / "fnde.py",
    ROOT / "modules" / "mapeamento_nova_planilha.py",
    PLANILHA,
]
missing = [str(p) for p in REQUIRED if not p.exists()]
if missing:
    raise SystemExit("Arquivos ausentes:\n" + "\n".join(missing))

configs = {}
for name in ["sistema.json", "codigos_ativos.json", "planilha_base.json"]:
    configs[name] = json.loads((ROOT / "config" / name).read_text(encoding="utf-8"))

plan_cfg = configs["planilha_base.json"]
if plan_cfg.get("arquivo") != "data/RREO-TCM+FNDE PLANILHA BASE.xlsx":
    raise SystemExit("config/planilha_base.json aponta para arquivo de planilha incorreto.")
if plan_cfg.get("aba") != ABA:
    raise SystemExit(f"config/planilha_base.json deve usar a aba {ABA!r}.")

wb = load_workbook(PLANILHA, read_only=True, data_only=False)
try:
    if ABA not in wb.sheetnames:
        raise SystemExit(f"Aba oficial ausente. Disponíveis: {', '.join(wb.sheetnames)}")
finally:
    wb.close()

sistema = configs["sistema.json"]
if sistema.get("rreo_chave_primaria") not in (None, "NOME_EXTERNO_UF"):
    raise SystemExit("RREO deve usar NOME_EXTERNO_UF como chave primária.")
if sistema.get("backup_drive") is True:
    raise SystemExit("backup_drive deve permanecer false.")

painel = (ROOT / "pages" / "1_Painel.py").read_text(encoding="utf-8")
if 'CHECKPOINT_SCHEMA_VERSION = "PLANILHA_SEQUENCIA_V2"' not in painel:
    raise SystemExit("Versão de esquema dos checkpoints não está protegida.")

print(f"OK: planilha oficial contém a aba {ABA!r}.")
print("OK: configuração da planilha sincronizada.")
print("OK: checkpoints versionados para impedir retomada de planilhas antigas.")
print("OK: estrutura do projeto validada.")
