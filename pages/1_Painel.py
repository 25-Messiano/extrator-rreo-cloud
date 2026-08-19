from __future__ import annotations

import hashlib
import random
import re
import shutil
import tempfile
import unicodedata
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

import streamlit as st
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from integrations.google_storage import (
    download_bytes,
    download_file,
    download_pdf,
    find_fnde_folder,
    find_rreo_folder,
    health_check,
    list_fnde_pdfs,
    list_pdfs,
    list_results,
    list_states,
    upload_result,
)
from modules.rreo import process as processar_rreo, identify_internal_municipality
from modules.fnde import process as processar_fnde
from modules.mapeamento_nova_planilha import (
    ABA_DESTINO,
    CAMPOS_FNDE_AUTORIZADOS,
    CAMPOS_RREO_AUTORIZADOS,
    carregar_municipios_da_planilha,
    obter_aba_destino,
    preencher_fnde_nova_planilha,
    preencher_rreo_nova_planilha,
    validar_estrutura_planilha,
)
from core.config_manager import load_json
from core.auditoria import timestamp, write_audit, write_missing
from core.auditoria_rreo import append_rreo_log
from core.auditoria_fnde import append_fnde_log
from core.indice_rreo import build_rreo_index, localizar_por_municipio
from core.indice_fnde import build_fnde_index
from core.indice_processamento import montar_plano
from core.politica_operacoes import politica_da_execucao
from core.processamento_lotes import (
    BatchSettings,
    checkpoint_filename,
    chunks,
    processed_codes_from_state,
    read_checkpoint_state,
    write_checkpoint_state,
)
from core.processamento_lotes_combinado import executar_lote_politica
from core.monitor_execucao import ExecutionMonitor, render_gauge_html
from core.controle_cancelamento import CancellationToken


st.markdown("""
<style>
.live-gauge-wrap{background:linear-gradient(180deg,#fff,#f8fafc);border:1px solid #dbe3ef;border-radius:16px;padding:10px 12px 8px;margin:8px 0 12px;box-shadow:0 5px 18px rgba(15,23,42,.06)}
.live-gauge-title{text-align:center;font-weight:800;font-size:.88rem;letter-spacing:.06em;color:#334155}.live-gauge{width:100%;max-height:245px;display:block;margin:auto}.gauge-mark{font:700 14px sans-serif;fill:#64748b}.gauge-value{font:800 24px sans-serif;fill:#0f172a}.gauge-unit{font:600 12px sans-serif;fill:#64748b}.gauge-stats{display:flex;justify-content:space-around;gap:8px;font-size:.82rem;color:#475569;margin-top:-4px}
</style>
""", unsafe_allow_html=True)

# ============================================================
# CONFIGURAÇÕES GERAIS
# ============================================================

BASE_DIR = Path(__file__).resolve().parents[1]

CODIGOS_CONFIG = load_json("codigos_ativos.json")
SISTEMA_CONFIG = load_json("sistema.json")
PERMITIR_RREO = bool(SISTEMA_CONFIG.get("processar_rreo", True))
PERMITIR_FNDE = bool(SISTEMA_CONFIG.get("processar_fnde", True))
PROCESSAR_RREO = PERMITIR_RREO
PROCESSAR_FNDE = PERMITIR_FNDE
GERAR_LOG_RREO = bool(SISTEMA_CONFIG.get("gerar_log_rreo", SISTEMA_CONFIG.get("gerar_log_processamento", True)))
GERAR_LOG_FNDE = bool(SISTEMA_CONFIG.get("gerar_log_fnde", SISTEMA_CONFIG.get("gerar_log_processamento", True)))
GERAR_NAO_ENCONTRADOS = bool(SISTEMA_CONFIG.get("gerar_municipios_nao_encontrados", True))
GERAR_AUDITORIA = bool(SISTEMA_CONFIG.get("gerar_auditoria", True))
PROGRAMAS_FNDE_ATIVOS = [
    k for k, v in CODIGOS_CONFIG.get("fnde", {}).items()
    if v and k in CAMPOS_FNDE_AUTORIZADOS
]
LOT_SETTINGS = BatchSettings.from_mapping(SISTEMA_CONFIG)
SALVAR_CHECKPOINT_CLOUD = bool(SISTEMA_CONFIG.get("salvar_checkpoint_cloud", True))
CHECKPOINT_SCHEMA_VERSION = "PLANILHA_SEQUENCIA_V2"


PLANILHA_BASE = (
    BASE_DIR
    / "data"
    / "RREO-TCM+FNDE PLANILHA BASE.xlsx"
)

TODOS_CODIGOS_RREO = [
    "1.1", "1.2", "1.3", "1.4", "2.1", "2.1.1", "2.1.2",
    "2.2", "2.3", "2.4", "2.5", "2.6", "6.1.1", "6.2", "6.2.1",
]
CODIGOS_RREO = [
    codigo for codigo in TODOS_CODIGOS_RREO
    if CODIGOS_CONFIG.get("rreo", {}).get(codigo, False)
    and codigo in CAMPOS_RREO_AUTORIZADOS
]

CODIGOS_UF = {
    "AC": "12",
    "AL": "27",
    "AP": "16",
    "AM": "13",
    "BA": "29",
    "CE": "23",
    "DF": "53",
    "ES": "32",
    "GO": "52",
    "MA": "21",
    "MT": "51",
    "MS": "50",
    "MG": "31",
    "PA": "15",
    "PB": "25",
    "PR": "41",
    "PE": "26",
    "PI": "22",
    "RJ": "33",
    "RN": "24",
    "RS": "43",
    "RO": "11",
    "RR": "14",
    "SC": "42",
    "SP": "35",
    "SE": "28",
    "TO": "17",
}
TODAS_UFS = tuple(sorted(CODIGOS_UF))

NOMES_UF = {
    "ACRE": "AC",
    "ALAGOAS": "AL",
    "AMAPA": "AP",
    "AMAZONAS": "AM",
    "BAHIA": "BA",
    "CEARA": "CE",
    "DISTRITO FEDERAL": "DF",
    "ESPIRITO SANTO": "ES",
    "GOIAS": "GO",
    "MARANHAO": "MA",
    "MATO GROSSO": "MT",
    "MATO GROSSO DO SUL": "MS",
    "MINAS GERAIS": "MG",
    "PARA": "PA",
    "PARAIBA": "PB",
    "PARANA": "PR",
    "PERNAMBUCO": "PE",
    "PIAUI": "PI",
    "RIO DE JANEIRO": "RJ",
    "RIO GRANDE DO NORTE": "RN",
    "RIO GRANDE DO SUL": "RS",
    "RONDONIA": "RO",
    "RORAIMA": "RR",
    "SANTA CATARINA": "SC",
    "SAO PAULO": "SP",
    "SERGIPE": "SE",
    "TOCANTINS": "TO",
}


# ============================================================
# NORMALIZAÇÃO
# ============================================================

def normalizar_texto(value: Any) -> str:
    if value is None:
        return ""

    text = str(value).strip()

    text = unicodedata.normalize(
        "NFKD",
        text,
    )

    text = "".join(
        character
        for character in text
        if not unicodedata.combining(character)
    )

    text = text.upper()

    text = re.sub(
        r"[^A-Z0-9]+",
        " ",
        text,
    )

    return re.sub(
        r"\s+",
        " ",
        text,
    ).strip()


def normalizar_codigo_ibge(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, float):
        value = int(value)

    return re.sub(
        r"\D",
        "",
        str(value),
    )


def extrair_uf(valor: str) -> str:
    texto = normalizar_texto(valor)

    partes = texto.split()

    for parte in reversed(partes):
        if parte in CODIGOS_UF:
            return parte

    for nome_estado, uf in NOMES_UF.items():
        if nome_estado in texto:
            return uf

    return ""


def extrair_nome_arquivo(nome_arquivo: str) -> str:
    nome = Path(nome_arquivo).stem

    nome = re.sub(
        r"(?i)^RREO[_\s-]*MUNICIPAL[_\s-]*\d{4}A?",
        "",
        nome,
    )

    nome = re.sub(
        r"(?i)^[_\s-]*RREO[_\s-]*",
        "",
        nome,
    )

    nome = re.sub(
        r"\s*-\s*[A-Za-z]{2}\s*$",
        "",
        nome,
    )

    return normalizar_texto(nome)


def similaridade(
    nome_a: str,
    nome_b: str,
) -> float:
    a = normalizar_texto(nome_a)
    b = normalizar_texto(nome_b)

    if not a or not b:
        return 0.0

    if a == b:
        return 1.0

    if a in b or b in a:
        return 0.97

    return SequenceMatcher(
        None,
        a,
        b,
    ).ratio()


# ============================================================
# LEITURA DA PLANILHA-BASE
# ============================================================

def localizar_coluna_por_cabecalho(
    worksheet: Worksheet,
    termos: list[str],
) -> int | None:
    termos_normalizados = [
        normalizar_texto(termo)
        for termo in termos
    ]

    limite_linhas = min(
        20,
        worksheet.max_row,
    )

    for row in range(
        1,
        limite_linhas + 1,
    ):
        for column in range(
            1,
            worksheet.max_column + 1,
        ):
            value = normalizar_texto(
                worksheet.cell(
                    row=row,
                    column=column,
                ).value
            )

            if not value:
                continue

            if any(
                termo in value
                for termo in termos_normalizados
            ):
                return column

    return None


def localizar_colunas_codigos(
    worksheet: Worksheet,
    codigos: list[str],
) -> dict[str, int]:
    colunas: dict[str, int] = {}

    limite_linhas = min(
        20,
        worksheet.max_row,
    )

    codigos_ordenados = sorted(
        codigos,
        key=len,
        reverse=True,
    )

    for row in range(
        1,
        limite_linhas + 1,
    ):
        for column in range(
            1,
            worksheet.max_column + 1,
        ):
            value = normalizar_texto(
                worksheet.cell(
                    row=row,
                    column=column,
                ).value
            )

            if not value:
                continue

            for codigo in codigos_ordenados:
                codigo_normalizado = normalizar_texto(
                    codigo
                )

                padroes_aceitos = (
                    f"{codigo_normalizado} ",
                    f"{codigo_normalizado}-",
                )

                if (
                    value == codigo_normalizado
                    or value.startswith(padroes_aceitos)
                ):
                    if codigo not in colunas:
                        colunas[codigo] = column

                    break

    return colunas


def escolher_aba_principal(
    workbook: Any,
) -> Worksheet:
    """Retorna exclusivamente a aba oficial da nova planilha."""
    worksheet = obter_aba_destino(workbook)
    validar_estrutura_planilha(worksheet)
    return worksheet


def carregar_municipios(
    worksheet: Worksheet,
    uf: str,
) -> list[dict[str, Any]]:
    """Carrega municípios pela matriz oficial: C=IBGE e D=Ente Federado."""
    return carregar_municipios_da_planilha(
        worksheet,
        uf,
        CODIGOS_UF[uf],
    )


# ============================================================
# IDENTIFICAÇÃO DO MUNICÍPIO
# ============================================================

def localizar_municipio_por_nome(
    nome_referencia: str,
    municipios: list[dict[str, Any]],
    limite_minimo: float = 0.72,
) -> tuple[dict[str, Any] | None, float]:
    melhor_municipio = None
    melhor_nota = 0.0

    for municipio in municipios:
        nota = similaridade(
            nome_referencia,
            municipio["nome"],
        )

        if nota > melhor_nota:
            melhor_nota = nota
            melhor_municipio = municipio

    if melhor_nota < limite_minimo:
        return None, melhor_nota

    return melhor_municipio, melhor_nota


def localizar_municipio_pdf(
    arquivo_pdf: dict[str, Any],
    texto_pdf: str,
    municipios: list[dict[str, Any]],
) -> tuple[dict[str, Any] | None, float, str]:
    """Identifica município SOMENTE pelo nome externo do PDF.

    ``texto_pdf`` é deliberadamente ignorado aqui. O conteúdo interno serve
    apenas à auditoria e nunca altera a associação do arquivo ao município.
    """
    del texto_pdf
    nome_arquivo = extrair_nome_arquivo(arquivo_pdf["name"])
    municipio, nota = localizar_municipio_por_nome(nome_arquivo, municipios)
    return municipio, nota, "nome externo do arquivo"


def identificar_municipio_rreo_no_conteudo(
    texto_pdf: str,
    municipios: list[dict[str, Any]],
    uf_esperada: str,
) -> tuple[dict[str, Any] | None, float, str]:
    municipio, confianca, metodo, _modelo, _tentativas = identify_internal_municipality(
        texto_pdf=texto_pdf,
        municipios=municipios,
        uf_esperada=uf_esperada,
        usar_gemini=False,
    )
    return municipio, confianca, metodo


def _indice_rreo_por_municipio(
    arquivos_pdf: list[dict[str, Any]],
) -> dict[str, dict[str, Any]]:
    indice: dict[str, dict[str, Any]] = {}
    for arquivo in arquivos_pdf:
        chave = extrair_nome_arquivo(arquivo["name"])
        if chave and chave not in indice:
            indice[chave] = arquivo
    return indice


def _arquivo_rreo_por_indice_interno(
    municipio: dict[str, Any],
    indice_interno: dict[str, Any] | None,
) -> tuple[dict[str, Any] | None, float]:
    """Localiza RREO exclusivamente pelo NOME EXTERNO do arquivo.

    Primeiro tenta correspondência exata do nome normalizado. Se não houver,
    aplica similaridade somente entre o nome oficial do município e os nomes
    externos indexados. O conteúdo interno do PDF nunca participa da escolha.
    """
    if not indice_interno:
        return None, 0.0

    item = localizar_por_municipio(
        indice_interno,
        municipio["nome"],
        municipio["uf"],
    )
    if item:
        nota = 1.0
    else:
        alvo = normalizar_texto(municipio["nome"])
        uf_alvo = str(municipio.get("uf") or "").upper()
        melhor_item = None
        melhor_nota = 0.0

        for candidato in (indice_interno or {}).get("por_nome_uf", {}).values():
            if str(candidato.get("uf") or "").upper() != uf_alvo:
                continue
            nome_externo = normalizar_texto(candidato.get("municipio_arquivo") or "")
            nota_atual = similaridade(alvo, nome_externo)
            if nota_atual > melhor_nota:
                melhor_item = candidato
                melhor_nota = nota_atual

        # Evita associação arriscada, mas tolera pequenas diferenças de grafia
        # no nome EXTERNO (acentos, hífens, abreviações e ruído de filename).
        if melhor_item is None or melhor_nota < 0.72:
            return None, melhor_nota
        item = melhor_item
        nota = melhor_nota

    return {
        "name": item.get("name", ""),
        "blob_name": item.get("blob_name", ""),
        "size": item.get("size", 0),
        "updated": item.get("updated", ""),
        "municipio_interno": "",
        "metodo_identificacao": "NOME_EXTERNO",
    }, nota


def localizar_pdf_municipio(
    municipio: dict[str, Any],
    arquivos_pdf: list[dict[str, Any]],
    indice: dict[str, dict[str, Any]] | None = None,
) -> tuple[dict[str, Any] | None, float]:
    chave = normalizar_texto(municipio["nome"])
    indice_local = indice if indice is not None else _indice_rreo_por_municipio(arquivos_pdf)

    if chave in indice_local:
        return indice_local[chave], 1.0

    melhor_arquivo = None
    melhor_nota = 0.0
    for nome_pdf, arquivo in indice_local.items():
        nota = similaridade(chave, nome_pdf)
        if nota > melhor_nota:
            melhor_nota = nota
            melhor_arquivo = arquivo

    if melhor_nota < 0.65:
        return None, melhor_nota
    return melhor_arquivo, melhor_nota



# ============================================================
# FNDE COMBINADO E AUDITORIA
# ============================================================

def _codigo_ibge_nome_fnde(name: str) -> str:
    match = re.search(r"(?<!\d)(\d{7})(?!\d)", name)
    return match.group(1) if match else ""


def _indice_fnde_por_ibge(uf: str, year: int) -> dict[str, dict[str, Any]]:
    if not PROCESSAR_FNDE:
        return {}
    folder = find_fnde_folder(uf, year)
    if not folder:
        return {}
    return build_fnde_index(list_fnde_pdfs(folder, year), uf).get("por_ibge", {})


def _processar_fnde_municipio(
    arquivo: dict[str, Any] | None,
    pasta_temporaria: Path,
) -> tuple[dict[str, float], list[str]]:
    if not arquivo:
        return {}, ["PDF FNDE não encontrado"]
    caminho = pasta_temporaria / f"fnde_{arquivo['name']}"
    download_pdf(arquivo["blob_name"], caminho)
    try:
        valores, _, resultado = processar_fnde(caminho, enviar_imagens=True)
        return valores, list(resultado.avisos)
    finally:
        caminho.unlink(missing_ok=True)


def _preencher_fnde(worksheet: Worksheet, row: int, valores: dict[str, float]) -> int:
    """Preenche somente V:Y (PNAE, PNATE, PDDE, QSE) na nova matriz."""
    filtrados = {
        programa: valores[programa]
        for programa in PROGRAMAS_FNDE_ATIVOS
        if programa in valores
    }
    return preencher_fnde_nova_planilha(worksheet, row, filtrados)


# ============================================================
# PROCESSAMENTO EM LOTES
# ============================================================

def _job_id(
    operacao: str,
    ano: int,
    uf_saida: str,
    modo_execucao: str,
    codigos_ibge: list[str] | tuple[str, ...] | set[str],
) -> str:
    """Cria uma chave exclusiva para checkpoint e retomada.

    Operações, modalidades, estados e seleções diferentes nunca compartilham
    checkpoint. A lista de municípios entra como resumo criptográfico curto,
    evitando nomes de arquivo enormes.
    """
    operation = normalizar_texto(operacao).replace(" ", "_")
    mode = normalizar_texto(modo_execucao).replace(" ", "_")
    scope = "BRASIL" if modo_execucao == "Todos os Estados" else uf_saida.upper()
    codes = "|".join(sorted(str(code) for code in codigos_ibge if str(code)))
    selection_hash = hashlib.sha1(codes.encode("utf-8")).hexdigest()[:12]
    return f"{CHECKPOINT_SCHEMA_VERSION}_{operation}_{mode}_{ano}_{scope}_{selection_hash}"


def _latest_checkpoint(job_id: str) -> dict[str, Any] | None:
    if not LOT_SETTINGS.resume_enabled:
        return None
    items = list_results(f"CHECKPOINTS/{job_id}")
    return items[0] if items else None


def _rreo_worker_payload(payload: dict[str, Any], temp_root: Path) -> dict[str, Any]:
    arquivo = payload.get("arquivo_rreo")
    if not arquivo:
        return {"values": {}, "text": "", "error": "PDF RREO não encontrado"}
    folder = temp_root / "rreo" / payload["codigo_ibge"]
    folder.mkdir(parents=True, exist_ok=True)
    values, text = processar_um_pdf(arquivo, folder)
    municipio_interno, confianca, origem, modelo, tentativas = identify_internal_municipality(
        texto_pdf=text,
        municipios=payload["trabalho"]["municipios"],
        uf_esperada=payload["uf"],
        usar_gemini=False,
    )
    return {
        "values": values,
        "text": text,
        "municipio_interno": municipio_interno,
        "confianca_municipio": confianca,
        "origem_municipio": origem,
        "modelo_municipio": modelo,
        "tentativas_municipio": tentativas,
        # A conferência interna é somente auditoria. Nunca bloqueia nem muda
        # o município escolhido pelo nome externo do arquivo.
        "error": "",
    }


def _fnde_worker_payload(payload: dict[str, Any], temp_root: Path) -> dict[str, Any]:
    arquivo = payload.get("arquivo_fnde")
    if not arquivo:
        return {"values": {}, "warnings": [], "result": None, "error": "PDF FNDE não encontrado"}
    folder = temp_root / "fnde" / payload["codigo_ibge"]
    folder.mkdir(parents=True, exist_ok=True)
    caminho = folder / arquivo["name"]
    download_pdf(arquivo["blob_name"], caminho)
    try:
        values, _, result = processar_fnde(caminho, enviar_imagens=True)
        return {
            "values": values,
            "warnings": list(result.avisos),
            "result": result,
            "error": "",
        }
    finally:
        caminho.unlink(missing_ok=True)


def _save_partial_result(path: Path, cloud_name: str = "Checkpoint local disponível") -> None:
    st.session_state["last_result"] = {
        "name": path.name,
        "bytes": path.read_bytes(),
        "cloud": cloud_name,
        "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }

# ============================================================
# PROCESSAMENTO
# ============================================================

def preencher_resultados(
    worksheet: Worksheet,
    row: int,
    resultados: dict[str, float | None],
    colunas_codigos: dict[str, int] | None = None,
) -> int:
    """Preenche apenas os destinos RREO autorizados pelo mapa fixo.

    ``colunas_codigos`` é mantido apenas por compatibilidade com checkpoints
    antigos do fluxo, mas não participa mais da decisão de destino.
    """
    del colunas_codigos
    return preencher_rreo_nova_planilha(worksheet, row, resultados)


def processar_um_pdf(
    arquivo_pdf: dict[str, Any],
    pasta_temporaria: Path,
) -> tuple[dict[str, float | None], str]:
    caminho_pdf = (
        pasta_temporaria
        / arquivo_pdf["name"]
    )

    download_pdf(
        blob_name=arquivo_pdf["blob_name"],
        destination=caminho_pdf,
    )

    resultados, texto = processar_rreo(
        caminho_pdf,
        CODIGOS_RREO,
    )

    try:
        caminho_pdf.unlink(
            missing_ok=True
        )
    except OSError:
        pass

    return resultados, texto


def gerar_nome_saida(
    uf: str,
    modo: str,
    municipio: dict[str, Any] | None = None,
    operacao: str = "RREO e FNDE",
) -> str:
    agora = datetime.now().strftime("%Y%m%d_%H%M%S")
    prefixo = {
        "RREO": "RREO",
        "FNDE": "FNDE",
        "RREO e FNDE": "RREO_FNDE",
    }.get(operacao, "RREO_FNDE")

    if modo == "Todos os Estados":
        return f"{prefixo}_TODOS_OS_ESTADOS_{agora}.xlsx"

    if modo == "Município único" and municipio:
        nome = normalizar_texto(municipio["nome"]).replace(" ", "_")
        return f"{prefixo}_{municipio['codigo_ibge']}_{nome}_{uf}_{agora}.xlsx"

    return f"{prefixo}_ESTADO_{uf}_{agora}.xlsx"


# ============================================================
# CACHE LEVE
# ============================================================

@st.cache_data(
    ttl=300,
    show_spinner=False,
)
def carregar_estados_cloud(
    year: int,
    include_rreo: bool,
    include_fnde: bool,
) -> list[str]:
    """Lista estados usando a API nova, com compatibilidade defensiva.

    A compatibilidade evita derrubar o painel caso o Render ainda esteja
    executando uma versão antiga de integrations/google_storage.py durante
    a troca de arquivos.
    """
    try:
        return list_states(
            year=year,
            include_rreo=include_rreo,
            include_fnde=include_fnde,
        )
    except TypeError as error:
        if "unexpected keyword argument" not in str(error):
            raise
        return list_states()


@st.cache_data(
    ttl=180,
    show_spinner=False,
)
def carregar_pdfs_estado(
    estado_cloud: str,
    year: int,
) -> list[dict[str, Any]]:
    try:
        return list_pdfs(
            estado_cloud,
            year=year,
        )
    except TypeError as error:
        if "unexpected keyword argument" not in str(error):
            raise
        return list_pdfs(estado_cloud)


@st.cache_data(ttl=1800, show_spinner=False)
def carregar_indice_rreo_interno(
    uf: str,
    year: int,
    arquivos_serializados: tuple[tuple[str, str, int, str], ...],
    municipios_serializados: tuple[tuple[int, str, str, str, str, str], ...],
) -> dict[str, Any]:
    arquivos = [
        {"name": name, "blob_name": blob, "size": size, "updated": updated}
        for name, blob, size, updated in arquivos_serializados
    ]
    municipios = [
        {
            "row": row, "codigo_ibge": code, "nome": name,
            "uf": city_uf, "ente": ente, "nome_normalizado": normalized,
        }
        for row, code, name, city_uf, ente, normalized in municipios_serializados
    ]
    return build_rreo_index(arquivos, uf=uf)


# ============================================================
# INTERFACE
# ============================================================
# ============================================================
# INTERFACE
# ============================================================


RUNTIME_STATE_KEYS = (
    "job",
    "logs",
    "last_result",
    "resultado_parcial",
    "checkpoint_carregado",
    "codigos_concluidos",
    "trabalho_atual",
    "execution_started",
)


def limpar_processamento_anterior() -> None:
    """Remove somente o estado transitório da execução anterior.

    Os valores atuais dos widgets não são removidos. Assim, mudar operação,
    modalidade, ano, estado ou municípios limpa progresso, logs, resultado e
    retomada antiga sem apagar a nova escolha do usuário.
    """
    for key in RUNTIME_STATE_KEYS:
        st.session_state.pop(key, None)

    # Índices de execução guardados na sessão devem ser isolados por contexto.
    for key in list(st.session_state):
        if str(key).startswith(("indice_rreo::", "indice_fnde::", "plano::")):
            st.session_state.pop(key, None)


def sincronizar_contexto_execucao(signature: str) -> None:
    previous = st.session_state.get("active_execution_context")
    if previous is not None and previous != signature:
        limpar_processamento_anterior()
    st.session_state["active_execution_context"] = signature


import pandas as pd
from ui.theme import apply_theme, metric_card, render_sidebar

st.set_page_config(page_title="Painel de Extração", page_icon="📊", layout="wide", initial_sidebar_state="expanded")
apply_theme()
render_sidebar()

status = health_check()
cloud_ok = bool(status.get("ok"))
gemini_ok = bool((__import__("os").getenv("GEMINI_API_KEY") or __import__("os").getenv("GOOGLE_API_KEY") or "").strip())

st.markdown(
    '<div class="hero-row"><div><div class="hero-title">Painel Único de Extração</div>'
    '<div class="hero-sub">RREO e FNDE no mesmo fluxo, com seleção de municípios, logs e auditoria.</div>'
    '</div><span class="online">● Sistema Online</span></div>',
    unsafe_allow_html=True,
)

abrangencias = [
    "Estado inteiro",
    "Município único",
    "Municípios selecionados",
    "Todos os Estados",
    "Amostra",
]

operacoes_habilitadas: list[str] = []
if PERMITIR_RREO:
    operacoes_habilitadas.extend(
        f"RREO — {abrangencia}" for abrangencia in abrangencias
    )
if PERMITIR_FNDE:
    operacoes_habilitadas.extend(
        f"FNDE — {abrangencia}" for abrangencia in abrangencias
    )
if PERMITIR_RREO and PERMITIR_FNDE:
    operacoes_habilitadas.extend(
        f"RREO + FNDE — {abrangencia}" for abrangencia in abrangencias
    )

if not operacoes_habilitadas:
    st.error("Nenhuma operação está habilitada em Configurações.")
    st.stop()

execucao_escolhida = st.selectbox(
    "Processo e abrangência",
    options=operacoes_habilitadas,
    index=0,
    key="execucao_escolhida_widget",
    on_change=limpar_processamento_anterior,
    help=(
        "Cada item é uma execução independente. RREO executa somente RREO; "
        "FNDE executa somente FNDE; RREO + FNDE executa os dois fluxos."
    ),
)

operacao_rotulo, modo = [
    parte.strip()
    for parte in execucao_escolhida.split("—", maxsplit=1)
]
operacao = (
    "RREO e FNDE"
    if operacao_rotulo == "RREO + FNDE"
    else operacao_rotulo
)

# A política é a única autoridade global sobre os módulos executados.
politica_execucao = politica_da_execucao(execucao_escolhida)
PROCESSAR_RREO = politica_execucao.usar_rreo
PROCESSAR_FNDE = politica_execucao.usar_fnde
todos_os_estados = politica_execucao.abrangencia.value == "Todos os Estados"

st.caption(
    f"Execução selecionada: {execucao_escolhida}. "
    f"RREO: {'SIM' if PROCESSAR_RREO else 'NÃO'} | "
    f"FNDE: {'SIM' if PROCESSAR_FNDE else 'NÃO'}"
)

with st.popover("🔎 O que o sistema está fazendo?", use_container_width=False):
    st.markdown(
        """
        **Recursos usados nesta execução**

        - **Python:** coordena o fluxo, identifica municípios, valida códigos IBGE e grava a planilha.
        - **Gemini Vision:** interpreta visualmente PDFs FNDE quando necessário.
        - **OCR (Tesseract):** lê textos e valores diretamente das imagens dos PDFs FNDE.
        - **PyPDF/PDFium:** abre os PDFs e converte as páginas em imagens para leitura visual.
        - **OpenPyXL:** preenche o Excel e atualiza LOG_RREO, LOG_FNDE, AUDITORIA e pendências.
        - **Google Cloud Storage:** localiza PDFs, salva checkpoints e publica resultados.
        - **Processamento em lotes:** limita tarefas simultâneas para acelerar sem sobrecarregar o Render.
        - **Checkpoint e retomada:** salva o progresso e permite continuar um trabalho interrompido.

        O painel mostra o andamento; a extração e as validações são executadas nos módulos internos.
        """
    )

m1,m2,m3,m4=st.columns(4)
with m1: metric_card("☁","Armazenamento","Cloud Storage","Conectado" if cloud_ok else "Verificar","blue")
with m2: metric_card("🗄","Banco de Dados","SQLite","Ativo","purple")
with m3: metric_card("🛡","Credenciais GCS","Configuradas" if cloud_ok else "Pendentes","OK" if cloud_ok else "Atenção","green")
with m4: metric_card("✦","Gemini API","Configurado" if gemini_ok else "Pendente","OK" if gemini_ok else "Atenção","blue")

if not cloud_ok:
    st.error("Não foi possível conectar ao Google Cloud Storage.")
    st.code(status.get("message","Erro desconhecido."))
    st.stop()
if PROCESSAR_FNDE and not gemini_ok and not SISTEMA_CONFIG.get("fnde_usar_ocr_fallback", True):
    st.error("FNDE selecionado, mas Gemini e OCR fallback não estão disponíveis.")
    st.stop()
if not PLANILHA_BASE.exists():
    st.error("A planilha-base não foi encontrada.")
    st.stop()

anos_referencia = list(range(2030, 2022, -1))
ano = st.selectbox(
    "Ano de Referência",
    anos_referencia,
    index=anos_referencia.index(2025),
    key="ano_referencia_widget",
    on_change=limpar_processamento_anterior,
)

try:
    estados_cloud = carregar_estados_cloud(
        ano,
        PROCESSAR_RREO,
        PROCESSAR_FNDE,
    )
except Exception as error:
    st.error("Não foi possível listar os estados do Cloud Storage.")
    st.exception(error)
    st.stop()
if not estados_cloud and not todos_os_estados:
    st.warning("Nenhuma pasta de estado foi encontrada para a operação e o ano selecionados.")
    st.stop()
if todos_os_estados and not estados_cloud:
    st.warning(
        "Nenhuma pasta foi listada no Cloud para o ano, mas a execução nacional "
        "continuará pelas 27 UFs oficiais e registrará as fontes ausentes."
    )

st.markdown('<div class="section-card"><div class="section-title"><span class="section-num">1.</span>Estado e Municípios</div>',unsafe_allow_html=True)
c1,c2,c3=st.columns([1.15,1.45,1.7])
with c1:
    st.text_input(
        "Processamento selecionado",
        value=execucao_escolhida,
        disabled=True,
        help="Definido na lista única no topo do painel.",
    )
with c2:
    if todos_os_estados:
        estado_cloud = st.selectbox(
            "Estado (UF) — referência visual",
            estados_cloud or list(TODAS_UFS),
            key="estado_cloud_widget",
            on_change=limpar_processamento_anterior,
        )
    else:
        estado_cloud=st.selectbox(
            "Estado (UF)",
            estados_cloud,
            key="estado_cloud_widget",
            on_change=limpar_processamento_anterior,
        )
uf=extrair_uf(estado_cloud) or (estado_cloud if estado_cloud in CODIGOS_UF else "")
if not uf:
    uf=st.selectbox("UF",sorted(CODIGOS_UF))

try:
    arquivos_pdf = (
        carregar_pdfs_estado(estado_cloud, ano)
        if PROCESSAR_RREO
        else []
    )
    indice_fnde_atual = (
        _indice_fnde_por_ibge(uf, ano)
        if PROCESSAR_FNDE
        else {}
    )
    workbook_consulta=load_workbook(PLANILHA_BASE,read_only=False,data_only=False)
    worksheet_consulta=escolher_aba_principal(workbook_consulta)
    municipios=carregar_municipios(worksheet_consulta,uf)
    workbook_consulta.close()
    # Índice leve: usa somente o nome externo e não abre nenhum PDF.
    indice_rreo_atual = build_rreo_index(arquivos_pdf, uf) if PROCESSAR_RREO else {}
except Exception as error:
    st.error("Não foi possível preparar os dados do estado.")
    st.exception(error)
    st.stop()

municipio_selecionado=None
municipios_selecionados=list(municipios)
arquivo_selecionado=None
arquivos_selecionados=list(arquivos_pdf)
selection_note="Todos os municípios do estado."

with c3:
    if todos_os_estados:
        st.selectbox("Município", ["Todos os municípios de todos os estados"], disabled=True)
    elif modo=="Município único":
        municipio_selecionado=st.selectbox(
            "Município",
            municipios,
            format_func=lambda i:f"{i['nome']} - {i['uf']}",
        )
        municipios_selecionados=[municipio_selecionado]
        selection_note="Um município selecionado manualmente."
    elif modo=="Municípios selecionados":
        codigos_escolhidos=st.multiselect(
            "Escolha os municípios",
            options=[m["codigo_ibge"] for m in municipios],
            format_func=lambda code: next(
                f"{m['nome']} - {m['uf']} ({m['codigo_ibge']})"
                for m in municipios if m["codigo_ibge"]==code
            ),
            default=[],
            help="Pesquise pelo nome ou código IBGE e marque quantos desejar.",
        )
        codigos_set=set(codigos_escolhidos)
        municipios_selecionados=[m for m in municipios if m["codigo_ibge"] in codigos_set]
        selection_note=f"{len(municipios_selecionados)} município(s) selecionado(s) manualmente."
    elif modo=="Amostra":
        amostra_cols=st.columns([1.0,1.4])
        with amostra_cols[0]:
            quantidade_amostra=st.number_input(
                "Quantidade",
                min_value=1,
                max_value=max(1,min(len(municipios),100)),
                value=min(5,max(1,len(municipios))),
                step=1,
            )
        with amostra_cols[1]:
            criterio_amostra=st.selectbox(
                "Critério",
                ["Primeiros da lista","Aleatórios (repetível)"],
            )
        quantidade=min(int(quantidade_amostra),len(municipios))
        if criterio_amostra=="Aleatórios (repetível)":
            municipios_selecionados=random.Random(f"{estado_cloud}|{ano}|RREO").sample(municipios,quantidade)
            municipios_selecionados=sorted(municipios_selecionados,key=lambda m:normalizar_texto(m["nome"]))
        else:
            municipios_selecionados=municipios[:quantidade]
        selection_note=f"Amostra de {len(municipios_selecionados)} município(s): {criterio_amostra}."
    else:
        st.selectbox("Município",["Todos os municípios"],disabled=True)

# A assinatura completa impede que uma execução antiga permaneça visível ou
# seja retomada quando o usuário muda operação, modalidade, ano, UF ou seleção.
context_codes = [m["codigo_ibge"] for m in municipios_selecionados]
context_signature = "|".join([
    execucao_escolhida,
    str(ano),
    "TODOS" if todos_os_estados else uf,
    ",".join(sorted(context_codes)),
])
sincronizar_contexto_execucao(context_signature)

st.markdown('</div>',unsafe_allow_html=True)

if modo=="Município único" and not todos_os_estados:
    arquivo_selecionado,nota_pdf=_arquivo_rreo_por_indice_interno(municipio_selecionado, indice_rreo_atual)
    nomes=[a["name"] for a in arquivos_pdf]
    if nomes:
        indice=nomes.index(arquivo_selecionado["name"]) if arquivo_selecionado else 0
        nome_escolhido=st.selectbox("PDF confirmado",nomes,index=indice)
        arquivo_selecionado=next(a for a in arquivos_pdf if a["name"]==nome_escolhido)
    arquivos_selecionados=[arquivo_selecionado] if arquivo_selecionado else []
elif not todos_os_estados and modo in {"Municípios selecionados","Amostra"}:
    arquivos_selecionados=[]
    for municipio_item in municipios_selecionados:
        arquivo_item,_=_arquivo_rreo_por_indice_interno(municipio_item, indice_rreo_atual)
        if arquivo_item and arquivo_item not in arquivos_selecionados:
            arquivos_selecionados.append(arquivo_item)

if not todos_os_estados and not municipios_selecionados:
    st.warning("Selecione pelo menos um município antes de processar.")

monitor = st.session_state.get("execution_monitor")
if not isinstance(monitor, ExecutionMonitor):
    monitor = ExecutionMonitor()
    st.session_state["execution_monitor"] = monitor
cancel_token = st.session_state.get("execution_cancel_token")
if not isinstance(cancel_token, CancellationToken):
    cancel_token = CancellationToken()
    st.session_state["execution_cancel_token"] = cancel_token

left,mid,right=st.columns([1.55,1.05,1.0])
with left:
    st.markdown('<div class="section-card"><div class="section-title"><span class="section-num">2.</span>Arquivos Encontrados</div>',unsafe_allow_html=True)
    a,b,c,d=st.columns(4)
    arquivos_resumo=arquivos_pdf if todos_os_estados or modo=="Estado inteiro" else arquivos_selecionados
    municipios_resumo=municipios if todos_os_estados or modo=="Estado inteiro" else municipios_selecionados
    codigos_resumo={m["codigo_ibge"] for m in municipios_resumo}
    fnde_resumo=[item for code,item in indice_fnde_atual.items() if code in codigos_resumo]
    total_size=(
        sum(int(x.get("size") or 0) for x in arquivos_resumo if x)
        + sum(int(x.get("size") or 0) for x in fnde_resumo if x)
    )
    total_pdfs=(len(arquivos_resumo) if PROCESSAR_RREO else 0) + (len(fnde_resumo) if PROCESSAR_FNDE else 0)
    with a: st.markdown(f'<div class="mini-stat"><div class="mini-label">Municípios selecionados</div><div class="mini-value">{len(municipios_resumo)}</div></div>',unsafe_allow_html=True)
    with b: st.markdown(f'<div class="mini-stat"><div class="mini-label">PDFs localizados</div><div class="mini-value">{total_pdfs}</div></div>',unsafe_allow_html=True)
    with c: st.markdown(f'<div class="mini-stat"><div class="mini-label">Tamanho Total</div><div class="mini-value">{total_size/1024/1024:.1f} MB</div></div>',unsafe_allow_html=True)
    with d: st.markdown(f'<div class="mini-stat"><div class="mini-label">Ano</div><div class="mini-value">{ano}</div></div>',unsafe_allow_html=True)
    preview=[]
    lista_preview=municipios if todos_os_estados or modo=="Estado inteiro" else municipios_selecionados
    for mun in lista_preview[:100]:
        arq,nota=_arquivo_rreo_por_indice_interno(mun, indice_rreo_atual)
        arq_fnde=indice_fnde_atual.get(mun["codigo_ibge"])
        preview.append({
            "Município":mun["nome"],
            "Código IBGE":mun["codigo_ibge"],
            "RREO":"Sim" if arq else "Não",
            "FNDE":"Sim" if arq_fnde else "Não",
            "Correspondência RREO":f"{nota*100:.0f}%" if arq else "-",
        })
    st.dataframe(pd.DataFrame(preview),use_container_width=True,hide_index=True,height=330)
    st.caption(f"Exibindo {min(100,len(preview))} de {len(lista_preview)} município(s) selecionado(s).")
    st.info(selection_note)
    st.markdown('</div>',unsafe_allow_html=True)

with mid:
    st.markdown('<div class="section-card"><div class="section-title"><span class="section-num">3.</span>Processamento ao vivo</div>',unsafe_allow_html=True)
    state=st.session_state.setdefault("job",{"status":"Aguardando","progress":0,"current":"Nenhum","success":0,"errors":0,"total":0})
    st.write("**Situação Atual**")
    st.info(monitor.status if monitor.status else state["status"])
    live_progress=st.empty()
    live_progress.progress(float(monitor.progress or state["progress"]),text=f"Progresso geral: {(monitor.progress or state['progress'])*100:.0f}%")
    gauge_slot=st.empty()
    gauge_slot.markdown(render_gauge_html(monitor),unsafe_allow_html=True)
    st.write("**Município / etapa atual**")
    live_current=st.empty()
    live_current.markdown(f"**{monitor.current}**  \n{monitor.stage} · Método: {monitor.method}")
    x,y,z=st.columns(3)
    x.metric("Concluídos",monitor.completed or state["success"])
    y.metric("Pendentes",max((monitor.total or state["total"])-(monitor.completed or state["success"])-monitor.errors,0))
    z.metric("Erros",monitor.errors or state["errors"])
    activity_slot=st.empty()
    activity_slot.code(monitor.recent_text(16),language=None)
    quantidade_selecionada=(len(municipios) if todos_os_estados or modo=="Estado inteiro" else len(municipios_selecionados))
    st.write("**Resumo antes de iniciar**")
    st.caption(f"Estado: {'TODOS' if todos_os_estados else uf} | Modo: {modo} | Municípios selecionados: {quantidade_selecionada} | Operação: {operacao} | RREO: {'SIM' if PROCESSAR_RREO else 'NÃO'} | FNDE: {'SIM' if PROCESSAR_FNDE else 'NÃO'}")
    cproc,ccancel=st.columns(2)
    with cproc:
        executar=st.button("▶ Processar agora",type="primary",use_container_width=True,disabled=(not todos_os_estados and not municipios_selecionados))
    with ccancel:
        cancelar=st.button("■ Cancelar processo",use_container_width=True,disabled=monitor.status not in {"Em andamento","Cancelamento solicitado"})
    if cancelar:
        cancel_token.request()
        monitor.update(status="Cancelamento solicitado",stage="Aguardando ponto seguro")
        monitor.event("WARNING","Cancelamento solicitado pelo usuário")
        st.warning("Cancelamento solicitado. O sistema salvará o último checkpoint seguro.")
    st.markdown('</div>',unsafe_allow_html=True)

with right:
    st.markdown('<div class="section-card"><div class="section-title"><span class="section-num">4.</span>Logs do Sistema ao vivo</div>',unsafe_allow_html=True)
    logbox=st.empty()
    logs=st.session_state.setdefault(
        "logs",
        [
            "Sistema pronto para iniciar.",
            f"RREO={len(arquivos_pdf)} PDF(s), FNDE={len(indice_fnde_atual)} PDF(s) em {uf}.",
        ],
    )
    logbox.code(monitor.recent_text(24),language=None)
    st.markdown('</div>',unsafe_allow_html=True)
    st.markdown('<div class="section-card"><div class="section-title"><span class="section-num">5.</span>Resultado</div>',unsafe_allow_html=True)
    result=st.session_state.get("last_result")
    if result:
        st.success(result["name"])
        st.download_button("⬇ Download",data=result["bytes"],file_name=result["name"],mime=result.get("mime","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),use_container_width=True)
        st.caption(result["cloud"])
    else:
        st.caption("O arquivo Excel aparecerá aqui após o processamento.")
    st.markdown('</div>',unsafe_allow_html=True)

if executar:
    cancel_token.clear()
    monitor.reset(0)
    monitor.event("INFO",f"Política ativa: RREO={'SIM' if PROCESSAR_RREO else 'NÃO'} | FNDE={'SIM' if PROCESSAR_FNDE else 'NÃO'}")
    monitor.event("INFO",f"Modo: {modo} | Estado: {'TODOS' if todos_os_estados else uf}")
    pasta_temporaria = Path(
        tempfile.mkdtemp(prefix="rreo_cloud_")
    )

    try:
        # Um único fluxo para os três modos:
        # 1) Município; 2) Estado; 3) Todos os Estados.
        # A diferença entre eles é apenas a lista de trabalhos montada abaixo.
        trabalhos: list[dict[str, Any]] = []

        if todos_os_estados:
            modo_execucao = "Todos os Estados"
            uf_saida = "BRASIL"
            municipio_saida = None

            workbook_nacional = load_workbook(
                PLANILHA_BASE,
                read_only=False,
                data_only=False,
            )
            worksheet_nacional = escolher_aba_principal(
                workbook_nacional
            )

            try:
                # Execução nacional é dirigida pelas 27 UFs oficiais da matriz,
                # não pela existência de pastas no Cloud. Assim nenhum estado é
                # omitido silenciosamente quando RREO ou FNDE estiver ausente.
                for uf_item in TODAS_UFS:
                    pasta_rreo = find_rreo_folder(uf_item, ano) if PROCESSAR_RREO else None
                    pasta_fnde = find_fnde_folder(uf_item, ano) if PROCESSAR_FNDE else None
                    estado_item = pasta_rreo or pasta_fnde or uf_item
                    try:
                        arquivos_item = carregar_pdfs_estado(uf_item, ano) if PROCESSAR_RREO else []
                    except Exception as error_estado:
                        arquivos_item = []
                        logs.append(
                            f"{datetime.now().strftime('%H:%M:%S')}  "
                            f"RREO {uf_item}: falha ao listar PDFs; continuando: {error_estado}"
                        )
                    municipios_item = carregar_municipios(
                        worksheet_nacional,
                        uf_item,
                    )
                    indice_interno_item = build_rreo_index(arquivos_item, uf_item) if PROCESSAR_RREO else {}
                    try:
                        indice_fnde_item = _indice_fnde_por_ibge(uf_item, ano) if PROCESSAR_FNDE else {}
                    except Exception as error_estado:
                        indice_fnde_item = {}
                        logs.append(
                            f"{datetime.now().strftime('%H:%M:%S')}  "
                            f"FNDE {uf_item}: falha ao listar PDFs; continuando: {error_estado}"
                        )
                    trabalhos.append(
                        {
                            "estado_cloud": estado_item,
                            "uf": uf_item,
                            "arquivos": arquivos_item,
                            "indice_rreo": indice_interno_item,
                            "indice_fnde": indice_fnde_item,
                            "municipios": municipios_item,
                            "pasta_rreo_encontrada": bool(pasta_rreo),
                            "pasta_fnde_encontrada": bool(pasta_fnde),
                        }
                    )
            finally:
                workbook_nacional.close()

        else:
            modo_execucao = modo
            uf_saida = uf
            municipio_saida = municipio_selecionado

            trabalhos.append(
                {
                    "estado_cloud": estado_cloud,
                    "uf": uf,
                    "arquivos": (
                        arquivos_pdf
                        if modo == "Estado inteiro"
                        else arquivos_selecionados
                    ),
                    "indice_rreo": indice_rreo_atual,
                    "indice_fnde": indice_fnde_atual,
                    "municipios": (
                        municipios
                        if modo == "Estado inteiro"
                        else municipios_selecionados
                    ),
                }
            )

        monitor.update(stage="Preparando planilha e fila",method="Python")
        monitor.event("INFO","Índices leves construídos pelos nomes externos")

        # A planilha-base é copiada e aberta uma única vez, independentemente
        # do modo escolhido. Assim, a estrutura de saída permanece a mesma.
        caminho_saida = pasta_temporaria / gerar_nome_saida(
            uf_saida,
            modo_execucao,
            municipio_saida,
            operacao,
        )

        shutil.copy2(
            PLANILHA_BASE,
            caminho_saida,
        )

        workbook = load_workbook(caminho_saida)
        worksheet = escolher_aba_principal(workbook)
        validar_estrutura_planilha(worksheet)
        colunas_codigos: dict[str, int] = {}  # compatibilidade; destino vem do módulo fixo

        # O Excel passa a existir como resultado parcial imediatamente. Se um
        # PDF ou serviço externo falhar depois, o usuário ainda terá o arquivo.
        workbook.save(caminho_saida)
        _save_partial_result(caminho_saida, "Planilha-base preparada; processamento em andamento.")

        total = sum(len(trabalho["municipios"]) for trabalho in trabalhos)

        monitor.total = total
        monitor.update(status="Em andamento",current="Preparando...",stage="Montando lotes",method="Python")
        state.update({
            "status": "Em andamento", "progress": 0, "current": "Preparando...",
            "success": 0, "errors": 0, "total": total,
        })

        divergencias: list[dict[str, Any]] = []
        missing_rows: list[dict[str, Any]] = []
        processed_ibge: set[str] = set()
        metrics = {
            "PDFs encontrados": sum(len(t["arquivos"]) + len(t.get("indice_fnde", {})) for t in trabalhos),
            "PDFs processados": 0,
            "Municípios preenchidos": 0,
            "Municípios pendentes": 0,
            "Campos preenchidos": 0,
            "Campos vazios": 0,
            "Erros críticos": 0,
            "Avisos": 0,
            "Upload Drive": "PENDENTE",
            "Operação": operacao,
            "RREO habilitado": "SIM" if PROCESSAR_RREO else "NÃO",
            "FNDE habilitado": "SIM" if PROCESSAR_FNDE else "NÃO",
            "Códigos RREO ativos": ", ".join(CODIGOS_RREO) if PROCESSAR_RREO else "",
            "Programas FNDE ativos": ", ".join(PROGRAMAS_FNDE_ATIVOS) if PROCESSAR_FNDE else "",
            "Data/hora de início": timestamp(),
        }
        progress_slot = st.progress(0, text="Iniciando processamento em lotes...")
        status_slot = st.empty()
        processados = 0

        # Constrói a fila nacional uma única vez. Os índices de arquivos já foram
        # montados por estado e são reutilizados durante todos os lotes.
        fila: list[dict[str, Any]] = []
        for trabalho in trabalhos:
            for municipio_item in trabalho["municipios"]:
                codigo = municipio_item["codigo_ibge"]
                arquivo_rreo, nota_rreo = _arquivo_rreo_por_indice_interno(
                    municipio_item,
                    trabalho.get("indice_rreo", {}),
                )
                fila.append({
                    "trabalho": trabalho,
                    "uf": trabalho["uf"],
                    "municipio": municipio_item,
                    "codigo_ibge": codigo,
                    "arquivo_rreo": arquivo_rreo,
                    "nota_rreo": nota_rreo,
                    "arquivo_fnde": trabalho.get("indice_fnde", {}).get(codigo),
                })

        job_id = _job_id(
            operacao,
            ano,
            uf_saida,
            modo_execucao,
            [item["codigo_ibge"] for item in fila],
        )
        checkpoint_state_folder = f"CHECKPOINTS/{job_id}"

        # Retomada automática: usa somente checkpoint da mesma operação,
        # modalidade, ano, escopo e conjunto de municípios.
        checkpoint = _latest_checkpoint(job_id)
        if checkpoint:
            checkpoint_local = pasta_temporaria / checkpoint["name"]
            try:
                download_file(checkpoint["blob_name"], checkpoint_local)

                # Valida o checkpoint ANTES de substituir a planilha-base atual.
                # Checkpoints de versões antigas podem possuir outra aba de destino.
                checkpoint_wb = load_workbook(checkpoint_local, read_only=False, data_only=False)
                try:
                    checkpoint_ws = escolher_aba_principal(checkpoint_wb)
                    validar_estrutura_planilha(checkpoint_ws)
                    control_state = read_checkpoint_state(checkpoint_wb)
                    checkpoint_job_id = str(control_state.get("job_id", "") or "")
                    if checkpoint_job_id and checkpoint_job_id != job_id:
                        raise ValueError(
                            "Checkpoint pertence a outra versão/contexto de processamento."
                        )
                finally:
                    checkpoint_wb.close()

                workbook.close()
                shutil.copy2(checkpoint_local, caminho_saida)
                workbook = load_workbook(caminho_saida)
                worksheet = escolher_aba_principal(workbook)
                validar_estrutura_planilha(worksheet)
                colunas_codigos = {}  # compatibilidade; destino vem do módulo fixo
                control_state = read_checkpoint_state(workbook)
                processed_ibge.update(processed_codes_from_state(control_state))
                state["success"] = int(control_state.get("sucessos", "0") or 0)
                state["errors"] = int(control_state.get("erros", "0") or 0)
                logs.append(
                    f"{datetime.now().strftime('%H:%M:%S')}  Retomando checkpoint compatível: "
                    f"{len(processed_ibge)} município(s) já concluído(s)."
                )
            except Exception as checkpoint_error:
                logs.append(
                    f"{datetime.now().strftime('%H:%M:%S')}  "
                    f"Checkpoint antigo/incompatível ignorado: {checkpoint_error}"
                )
                # Restaura explicitamente a planilha-base oficial. Nesta etapa ainda
                # não houve gravação de municípios, portanto é seguro descartar o
                # checkpoint incompatível e iniciar uma execução nova.
                try:
                    workbook.close()
                except Exception:
                    pass
                shutil.copy2(PLANILHA_BASE, caminho_saida)
                workbook = load_workbook(caminho_saida)
                worksheet = escolher_aba_principal(workbook)
                validar_estrutura_planilha(worksheet)
                colunas_codigos = {}

        fila_pendente = [item for item in fila if item["codigo_ibge"] not in processed_ibge]
        total = len(fila)
        processados = len(processed_ibge)
        state["total"] = total
        state["progress"] = processados / total if total else 1.0

        cancelado=False
        for numero_lote, lote in enumerate(chunks(fila_pendente, LOT_SETTINGS.batch_size), start=1):
            if cancel_token.requested:
                cancelado=True
                monitor.event("WARNING","Cancelamento reconhecido antes do próximo lote")
                break
            monitor.update(stage=f"Lote {numero_lote}",method="Paralelismo limitado")
            monitor.event("INFO",f"Lote {numero_lote} iniciado com {len(lote)} município(s)")
            status_slot.info(
                f"Lote {numero_lote}: {len(lote)} município(s) | "
                f"RREO={LOT_SETTINGS.rreo_workers} worker(s) | "
                f"FNDE={LOT_SETTINGS.fnde_workers} worker(s) | Gemini=1 chamada"
            )

            rreo_results, fnde_results = executar_lote_politica(
                politica_execucao,
                lote,
                (lambda item: _rreo_worker_payload(item, pasta_temporaria)) if PROCESSAR_RREO else None,
                (lambda item: _fnde_worker_payload(item, pasta_temporaria)) if PROCESSAR_FNDE else None,
                LOT_SETTINGS.rreo_workers,
                LOT_SETTINGS.fnde_workers,
                LOT_SETTINGS.batch_timeout_seconds,
            )

            # A escrita no Excel ocorre somente na thread principal. Isso evita
            # corrupção da planilha mesmo com a extração paralela.
            for index, payload in enumerate(lote):
                municipio_encontrado = payload["municipio"]
                trabalho = payload["trabalho"]
                uf_item = payload["uf"]
                codigo_ibge = payload["codigo_ibge"]
                arquivo_rreo = payload["arquivo_rreo"]
                arquivo_fnde = payload["arquivo_fnde"]
                nota_rreo = payload["nota_rreo"]
                identificacao = f"{codigo_ibge} - {municipio_encontrado['nome']}/{uf_item}"
                state["current"] = identificacao
                monitor.update(current=identificacao,stage="Extraindo RREO/FNDE",method="Python/OCR/Gemini")
                monitor.event("INFO",f"Iniciando {identificacao}")
                live_current.markdown(f"**{identificacao}**  \n{monitor.stage} · Método: {monitor.method}")

                rreo_values: dict[str, float | None] = {}
                fnde_values: dict[str, float] = {}
                fnde_warnings: list[str] = []
                erros_municipio: list[str] = []
                rreo_count = 0
                fnde_count = 0

                rreo_data, rreo_error = rreo_results.get(index, (None, None))
                fnde_data, fnde_error = fnde_results.get(index, (None, None))

                municipio_rreo = municipio_encontrado
                divergencia_nome_rreo = ""
                origem_municipio_rreo = "NOME_EXTERNO"
                confianca_municipio_rreo = float(nota_rreo or 1.0)
                acao_divergencia_rreo = "PROCESSADO_PELO_NOME_EXTERNO"
                municipio_interno_nome = ""

                if PROCESSAR_RREO:
                    if rreo_error:
                        erros_municipio.append(f"RREO: {type(rreo_error).__name__}: {rreo_error}")
                    elif rreo_data:
                        rreo_values = rreo_data.get("values", {})
                        municipio_interno = rreo_data.get("municipio_interno")
                        origem_auditoria = rreo_data.get("origem_municipio", "")
                        confianca_auditoria = float(rreo_data.get("confianca_municipio", 0.0) or 0.0)
                        nome_arquivo_rreo = extrair_nome_arquivo(arquivo_rreo["name"]) if arquivo_rreo else ""
                        if municipio_interno:
                            municipio_interno_nome = municipio_interno.get("nome", "")
                            if normalizar_texto(nome_arquivo_rreo) != normalizar_texto(municipio_interno_nome):
                                divergencia_nome_rreo = (
                                    f"NOME_ARQUIVO_DIVERGE_DO_CONTEUDO: arquivo={nome_arquivo_rreo}; "
                                    f"conteúdo={municipio_interno_nome}"
                                )
                                acao_divergencia_rreo = "PROCESSADO_PELO_NOME_EXTERNO; DIVERGENCIA_APENAS_NO_LOG"
                            else:
                                # Conteúdo interno confirmado, mas não muda origem, confiança
                                # nem destino: o nome externo continua sendo a regra oficial.
                                pass

                        # REGRA OFICIAL: o nome externo manda. A leitura interna serve
                        # somente para auditoria e jamais impede o preenchimento.
                        rreo_count = preencher_resultados(
                            worksheet,
                            municipio_encontrado["row"],
                            rreo_values,
                            colunas_codigos,
                        )
                        if divergencia_nome_rreo:
                            monitor.event("WARNING", f"RREO divergente (somente auditoria): {divergencia_nome_rreo}")
                        if rreo_data.get("error") and not municipio_interno:
                            # Falha na conferência interna não bloqueia o processamento pelo nome externo.
                            pass
                        if arquivo_rreo:
                            metrics["PDFs processados"] += 1

                fnde_result_obj = None
                if PROCESSAR_FNDE:
                    if fnde_error:
                        erros_municipio.append(f"FNDE: {type(fnde_error).__name__}: {fnde_error}")
                    elif fnde_data:
                        fnde_values = fnde_data.get("values", {})
                        fnde_warnings = fnde_data.get("warnings", [])
                        fnde_result_obj = fnde_data.get("result")
                        if fnde_data.get("error"):
                            erros_municipio.append(fnde_data["error"])
                        fnde_count = _preencher_fnde(
                            worksheet, municipio_encontrado["row"], fnde_values
                        )
                        if arquivo_fnde:
                            metrics["PDFs processados"] += 1

                filled = rreo_count + fnde_count
                expected = (len(CODIGOS_RREO) if PROCESSAR_RREO else 0) + (
                    len(PROGRAMAS_FNDE_ATIVOS) if PROCESSAR_FNDE else 0
                )
                metrics["Campos preenchidos"] += filled
                metrics["Campos vazios"] += max(expected - filled, 0)
                metrics["Avisos"] += len(fnde_warnings)

                if GERAR_LOG_RREO and PROCESSAR_RREO:
                    found_codes = [c for c, v in rreo_values.items() if v is not None]
                    missing_codes = [c for c in CODIGOS_RREO if rreo_values.get(c) is None]
                    append_rreo_log(workbook, {
                        "Data/Hora": timestamp(),
                        "PDF RREO": arquivo_rreo["name"] if arquivo_rreo else "",
                        "Município": municipio_encontrado["nome"],
                        "Código IBGE": codigo_ibge,
                        "UF": uf_item, "Ano": ano,
                        "Estado/Lote": trabalho["estado_cloud"],
                        "Linha da planilha": municipio_encontrado["row"],
                        "Município no nome do arquivo": extrair_nome_arquivo(arquivo_rreo["name"]) if arquivo_rreo else "",
                        "Município no conteúdo": municipio_interno_nome if divergencia_nome_rreo else "",
                        "Divergência de município": "SIM" if divergencia_nome_rreo else "NÃO",
                        "Ação adotada": acao_divergencia_rreo,
                        "Origem da identificação": origem_municipio_rreo,
                        "Confiança da identificação": f"{confianca_municipio_rreo:.2%}" if confianca_municipio_rreo else "",
                        "Campos RREO preenchidos": rreo_count,
                        "Códigos encontrados": ", ".join(found_codes),
                        "Códigos ausentes": ", ".join(missing_codes),
                        "Status": "OK" if rreo_count else ("ERRO" if rreo_error else "PARCIAL"),
                        "Erro resumido": "; ".join(erros_municipio),
                        "Observações": divergencia_nome_rreo,
                        "Método de extração": f"PARALELO_LIMITADO + {origem_municipio_rreo}",
                    })

                if GERAR_LOG_FNDE and PROCESSAR_FNDE:
                    found_fnde = [k for k, v in fnde_values.items() if float(v or 0.0) != 0.0]
                    missing_fnde = [k for k in PROGRAMAS_FNDE_ATIVOS if k not in found_fnde]
                    append_fnde_log(workbook, {
                        "Data/Hora": timestamp(),
                        "PDF FNDE": arquivo_fnde["name"] if arquivo_fnde else "",
                        "Município": municipio_encontrado["nome"],
                        "Código IBGE": codigo_ibge, "UF": uf_item, "Ano": ano,
                        "Estado/Lote": trabalho["estado_cloud"],
                        "Linha da planilha": municipio_encontrado["row"],
                        "Campos FNDE preenchidos": fnde_count,
                        "Programas encontrados": ", ".join(found_fnde),
                        "Programas ausentes": ", ".join(missing_fnde),
                        "Valores extraídos": "; ".join(f"{k}={v:.2f}" for k, v in fnde_values.items()),
                        "Status": "OK" if fnde_count else ("ERRO" if fnde_error else "PARCIAL"),
                        "Avisos": "; ".join(fnde_warnings),
                        "Erro resumido": "; ".join(erros_municipio),
                        "Método": getattr(fnde_result_obj, "metodo", "") if fnde_result_obj else "",
                        "Modelo Gemini": getattr(fnde_result_obj, "modelo", "") if fnde_result_obj else "",
                        "Tentativas": getattr(fnde_result_obj, "tentativas", 0) if fnde_result_obj else 0,
                    })

                if filled > 0:
                    state["success"] += 1
                    metrics["Municípios preenchidos"] += 1
                    processed_ibge.add(codigo_ibge)
                else:
                    state["errors"] += 1
                    metrics["Erros críticos"] += 1

                if erros_municipio:
                    divergencias.append({
                        "UF": uf_item, "Código IBGE": codigo_ibge,
                        "Município": municipio_encontrado["nome"],
                        "problema": "; ".join(erros_municipio),
                        "nota": nota_rreo if arquivo_rreo else 0,
                    })

                processados += 1
                monitor.complete(identificacao,error=(filled==0))
                state["progress"] = processados / total if total else 1.0
                gauge_slot.markdown(render_gauge_html(monitor),unsafe_allow_html=True)
                live_progress.progress(monitor.progress,text=f"Progresso geral: {monitor.progress*100:.0f}%")
                live_current.markdown(f"**{monitor.current}**  \n{monitor.stage} · Método: {monitor.method}")
                activity_slot.code(monitor.recent_text(16),language=None)
                logbox.code(monitor.recent_text(24),language=None)
                if cancel_token.requested:
                    cancelado=True
                    monitor.event("WARNING","Cancelamento reconhecido após o município atual")
                    break
                if processados % LOT_SETTINGS.ui_update_interval == 0 or processados == total:
                    progress_slot.progress(
                        state["progress"],
                        text=f"Progresso geral: {state['progress'] * 100:.0f}%",
                    )
                    logs.append(
                        f"{datetime.now().strftime('%H:%M:%S')}  {identificacao}: "
                        f"RREO={rreo_count}, FNDE={fnde_count}"
                    )
                    logbox.code("\n".join(logs[-12:]), language=None)

            # Checkpoint ao final de cada lote. A cópia local fica imediatamente
            # disponível para download, antes mesmo do upload ao Cloud.
            write_checkpoint_state(workbook, {
                "job_id": job_id,
                "operacao": operacao,
                "ano": ano,
                "codigos_processados": ",".join(sorted(processed_ibge)),
                "sucessos": state["success"],
                "erros": state["errors"],
                "total": total,
                "ultimo_lote": numero_lote,
                "atualizado_em": timestamp(),
            })
            workbook.save(caminho_saida)
            _save_partial_result(caminho_saida)

            if SALVAR_CHECKPOINT_CLOUD:
                checkpoint_path = pasta_temporaria / checkpoint_filename(job_id, processados)
                shutil.copy2(caminho_saida, checkpoint_path)
                try:
                    checkpoint_cloud = upload_result(checkpoint_path, checkpoint_state_folder)
                    st.session_state["last_result"]["cloud"] = checkpoint_cloud["blob_name"]
                except Exception as checkpoint_error:
                    logs.append(
                        f"{datetime.now().strftime('%H:%M:%S')}  "
                        f"Checkpoint local salvo; upload falhou: {checkpoint_error}"
                    )
            monitor.event("INFO",f"Checkpoint do lote {numero_lote} salvo")
            if cancelado:
                break

        if cancelado:
            workbook.save(caminho_saida)
            workbook.close()
            _save_partial_result(caminho_saida,"Processo cancelado; resultado parcial preservado.")
            state["status"]="Cancelado"
            state["current"]="Cancelado pelo usuário"
            monitor.update(status="Cancelado",current="Processo cancelado",stage="Checkpoint salvo",method="-")
            monitor.event("WARNING","Processo cancelado com segurança; resultado parcial disponível")
            st.warning("Processo cancelado. O último resultado parcial foi preservado.")
            st.rerun()

        if GERAR_NAO_ENCONTRADOS:
            for trabalho in trabalhos:
                for municipio in trabalho["municipios"]:
                    if municipio["codigo_ibge"] in processed_ibge:
                        continue
                    arquivo, _ = _arquivo_rreo_por_indice_interno(
                        municipio,
                        trabalho.get("indice_rreo", {}),
                    )
                    arquivo_fnde_pendente = trabalho.get("indice_fnde", {}).get(
                        municipio["codigo_ibge"]
                    )
                    faltas = []
                    if PROCESSAR_RREO and not arquivo:
                        faltas.append("RREO")
                    if PROCESSAR_FNDE and not arquivo_fnde_pendente:
                        faltas.append("FNDE")
                    missing_rows.append({
                        "Estado/UF": trabalho["uf"],
                        "Código IBGE": municipio["codigo_ibge"],
                        "Município da planilha-base": municipio["nome"],
                        "Situação": "SEM PDF FORNECIDO" if faltas else "PDF FORNECIDO MAS NÃO LIDO",
                        "PDF RREO correspondente": arquivo["name"] if arquivo else "",
                        "PDF FNDE correspondente": arquivo_fnde_pendente["name"] if arquivo_fnde_pendente else "",
                        "Observação": ("Faltando: " + ", ".join(faltas)) if faltas else "Não houve preenchimento seguro durante esta execução.",
                    })
            write_missing(workbook, missing_rows)
            metrics["Municípios pendentes"] = len(missing_rows)

        metrics["Data/hora de fim"] = timestamp()
        metrics["Arquivo de saída"] = caminho_saida.name
        if GERAR_AUDITORIA:
            write_audit(workbook, metrics)

        workbook.save(caminho_saida)
        _save_partial_result(caminho_saida, "Arquivo final salvo localmente; enviando ao Cloud...")

        cloud_message = "UPLOAD_FALHOU - download local disponível"
        try:
            resultado_cloud = upload_result(caminho_saida, uf_saida)
            metrics["Upload Drive"] = "OK"
            cloud_message = resultado_cloud["blob_name"]
        except Exception as upload_error:
            metrics["Upload Drive"] = "FALHOU"
            logs.append(
                f"{datetime.now().strftime('%H:%M:%S')}  "
                f"Upload final falhou; download local preservado: {upload_error}"
            )

        # Atualiza a auditoria com o resultado real do upload e salva novamente.
        if GERAR_AUDITORIA:
            write_audit(workbook, metrics)
        workbook.save(caminho_saida)
        workbook.close()
        _save_partial_result(caminho_saida, cloud_message)

        if metrics["Upload Drive"] == "OK":
            try:
                resultado_cloud = upload_result(caminho_saida, uf_saida)
                st.session_state["last_result"]["cloud"] = resultado_cloud["blob_name"]
            except Exception as upload_error:
                logs.append(f"Reenvio da auditoria falhou: {upload_error}")

        monitor.update(status="Concluído",current="Finalizado",stage="Resultado disponível",method="-")
        monitor.event("OK","Processamento concluído")
        state["status"] = "Concluído"
        state["current"] = "Finalizado"
        state["progress"] = 1.0

        logs.append(
            f"{datetime.now().strftime('%H:%M:%S')}  "
            f"Extração concluída: {state['success']} sucesso(s), "
            f"{state['errors']} erro(s)."
        )
        if metrics["Upload Drive"] == "OK":
            status_slot.success("Processamento concluído e Excel salvo no Cloud Storage.")
        else:
            status_slot.warning(
                "Processamento concluído. O upload falhou, mas o download local está disponível."
            )

        if divergencias:
            st.warning("Alguns arquivos precisam de conferência.")
            st.dataframe(
                divergencias,
                use_container_width=True,
                hide_index=True,
            )

        st.rerun()

    except Exception as error:
        monitor.update(status="Falha",stage="Erro",method="-")
        monitor.event("ERROR",str(error))
        state["status"] = "Falha"
        logs.append(f"ERRO: {error}")
        st.exception(error)

    finally:
        shutil.rmtree(
            pasta_temporaria,
            ignore_errors=True,
        )

st.markdown('<div class="section-card"><div class="section-title">Fluxo Operacional</div><div class="flow"><div class="flow-step"><div class="flow-num">1</div><div><div class="flow-name">Listagem</div><div class="flow-desc">PDFs localizados no Google Cloud Storage</div></div></div><div class="flow-arrow">→</div><div class="flow-step"><div class="flow-num">2</div><div><div class="flow-name">Extração</div><div class="flow-desc">Gemini lê Receitas Realizadas Até o Bimestre (b)</div></div></div><div class="flow-arrow">→</div><div class="flow-step"><div class="flow-num">3</div><div><div class="flow-name">Geração</div><div class="flow-desc">Excel preenchido automaticamente</div></div></div><div class="flow-arrow">→</div><div class="flow-step"><div class="flow-num">4</div><div><div class="flow-name">Upload</div><div class="flow-desc">Resultado salvo e liberado</div></div></div></div></div>',unsafe_allow_html=True)
st.markdown('<div class="footerbar">● Sistema operando com Gemini &nbsp;•&nbsp; Extração inteligente &nbsp;•&nbsp; Cache de listagem &nbsp;•&nbsp; Armazenamento seguro na nuvem</div>',unsafe_allow_html=True)
