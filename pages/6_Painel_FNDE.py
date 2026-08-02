from __future__ import annotations

import os
import re
import tempfile
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st

from core.config_manager import load_json
from openpyxl import load_workbook

from integrations.google_storage import (
    BUCKET_NAME,
    download_file,
    download_pdf,
    get_storage_client,
    health_check,
    list_results,
    upload_file,
)
from modules.fnde import COLUNAS_FNDE, normalizar_codigo_ibge, process
from core.auditoria import append_fnde_log, timestamp, write_audit, write_missing
from ui.theme import apply_theme, metric_card, render_sidebar


# ============================================================
# CONFIGURACOES
# ============================================================

BASE_DIR = Path(__file__).resolve().parents[1]
SISTEMA_CONFIG = load_json("sistema.json")
PERMITIR_TODOS_ESTADOS = bool(SISTEMA_CONFIG.get("fnde_todos_estados", False))
GERAR_LOG_FNDE = bool(SISTEMA_CONFIG.get("gerar_log_fnde", True))
GERAR_NAO_ENCONTRADOS = bool(SISTEMA_CONFIG.get("gerar_municipios_nao_encontrados", True))
GERAR_AUDITORIA = bool(SISTEMA_CONFIG.get("gerar_auditoria", True))
SALVAR_INCREMENTAL = bool(SISTEMA_CONFIG.get("fnde_salvar_incremental", True))
CHECKPOINT_INTERVALO = max(int(SISTEMA_CONFIG.get("fnde_checkpoint_intervalo", 10)), 1)
PLANILHA_BASE = BASE_DIR / "data" / "RREO-TCM+FNDE PLANILHA BASE.xlsx"

FNDE_ROOT_PREFIX = "ARQUIVO_DE_ESTADOS_RREO/"

FNDE_RESULTADOS_PREFIX = (
    "ARQUIVO_DE_ESTADOS_RREO/"
    "PLANILHAS_PROCESSADAS_FNDE/"
)

PADRAO_PASTA_FNDE = re.compile(
    r"^.+?\s+FNDE\s*-\s*([A-Z]{2})_(\d{4})$",
    re.IGNORECASE,
)

CODIGOS_UF = {
    "AC": "12", "AL": "27", "AP": "16", "AM": "13", "BA": "29",
    "CE": "23", "DF": "53", "ES": "32", "GO": "52", "MA": "21",
    "MT": "51", "MS": "50", "MG": "31", "PA": "15", "PB": "25",
    "PR": "41", "PE": "26", "PI": "22", "RJ": "33", "RN": "24",
    "RS": "43", "RO": "11", "RR": "14", "SC": "42", "SP": "35",
    "SE": "28", "TO": "17",
}

NOMES_UF = {
    "ACRE": "AC", "ALAGOAS": "AL", "AMAPA": "AP", "AMAZONAS": "AM",
    "BAHIA": "BA", "CEARA": "CE", "DISTRITO FEDERAL": "DF",
    "ESPIRITO SANTO": "ES", "GOIAS": "GO", "MARANHAO": "MA",
    "MATO GROSSO": "MT", "MATO GROSSO DO SUL": "MS",
    "MINAS GERAIS": "MG", "PARA": "PA", "PARAIBA": "PB",
    "PARANA": "PR", "PERNAMBUCO": "PE", "PIAUI": "PI",
    "RIO DE JANEIRO": "RJ", "RIO GRANDE DO NORTE": "RN",
    "RIO GRANDE DO SUL": "RS", "RONDONIA": "RO", "RORAIMA": "RR",
    "SANTA CATARINA": "SC", "SAO PAULO": "SP", "SERGIPE": "SE",
    "TOCANTINS": "TO",
}


# ============================================================
# FUNCOES DE APOIO
# ============================================================

def normalizar_texto(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or "").strip())
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^A-Z0-9]+", " ", text.upper())
    return re.sub(r"\s+", " ", text).strip()


def extrair_uf(value: str) -> str:
    text = normalizar_texto(value)

    for part in reversed(text.split()):
        if part in CODIGOS_UF:
            return part

    for state_name, uf in NOMES_UF.items():
        if state_name in text:
            return uf

    return ""


def dados_pasta_fnde(folder_name: str) -> tuple[str, int | None]:
    match = PADRAO_PASTA_FNDE.match(folder_name.strip())
    if not match:
        return extrair_uf(folder_name), None

    return match.group(1).upper(), int(match.group(2))


def pasta_rreo_correspondente(folder_name: str) -> str:
    text = re.sub(
        r"\s+FNDE\s*-\s*[A-Z]{2}_\d{4}\s*$",
        "",
        folder_name,
        flags=re.IGNORECASE,
    ).strip()
    uf, _ = dados_pasta_fnde(folder_name)
    return f"{text} - {uf}" if text and uf else text or uf


def listar_estados_fnde() -> list[str]:
    client = get_storage_client()
    iterator = client.list_blobs(
        BUCKET_NAME,
        prefix=FNDE_ROOT_PREFIX,
        delimiter="/",
    )
    list(iterator)

    folders: list[str] = []
    for prefix in iterator.prefixes:
        folder_name = prefix.rstrip("/").split("/")[-1]
        if PADRAO_PASTA_FNDE.match(folder_name):
            folders.append(folder_name)

    return sorted(folders, key=normalizar_texto)


def listar_pdfs_fnde(state: str) -> list[dict[str, Any]]:
    client = get_storage_client()
    prefix = f"{FNDE_ROOT_PREFIX}{state}/"
    files: list[dict[str, Any]] = []

    for blob in client.list_blobs(BUCKET_NAME, prefix=prefix):
        if not blob.name.lower().endswith(".pdf"):
            continue

        files.append(
            {
                "name": Path(blob.name).name,
                "blob_name": blob.name,
                "size": blob.size or 0,
                "updated": blob.updated,
            }
        )

    return sorted(files, key=lambda item: normalizar_texto(item["name"]))


def localizar_aba_principal(workbook: Any):
    best = workbook.active
    best_score = -1

    for worksheet in workbook.worksheets:
        score = 0
        for row in range(1, min(20, worksheet.max_row) + 1):
            for column in range(1, worksheet.max_column + 1):
                value = normalizar_texto(worksheet.cell(row, column).value)
                if "CODIGO IBGE" in value:
                    score += 10
                if "ENTE FEDERADO" in value or "MUNICIPIO" in value:
                    score += 10
                if "PNAE" in value:
                    score += 2
                if "PNATE" in value:
                    score += 2
                if "PDDE" in value:
                    score += 2
                if "QSE" in value or "QESE" in value:
                    score += 2

        if score > best_score:
            best = worksheet
            best_score = score

    return best


def localizar_coluna_ibge(worksheet: Any) -> int:
    for row in range(1, min(20, worksheet.max_row) + 1):
        for column in range(1, worksheet.max_column + 1):
            value = normalizar_texto(worksheet.cell(row, column).value)
            if "CODIGO IBGE" in value:
                return column

    raise RuntimeError("A coluna Codigo IBGE nao foi encontrada na planilha.")


def indice_linhas_ibge(worksheet: Any) -> dict[str, int]:
    column = localizar_coluna_ibge(worksheet)
    index: dict[str, int] = {}

    for row in range(1, worksheet.max_row + 1):
        code = normalizar_codigo_ibge(worksheet.cell(row, column).value)
        if len(code) == 7:
            index[code] = row

    return index


def preencher_linha(
    worksheet: Any,
    row: int,
    values: dict[str, float],
) -> None:
    for program, column in COLUNAS_FNDE.items():
        cell = worksheet.cell(row=row, column=column)
        cell.value = float(values.get(program, 0.0) or 0.0)
        cell.number_format = '#,##0.00'


def baixar_planilha_origem(
    state: str,
    uf: str,
    destination: Path,
) -> tuple[Path, str]:
    search_keys = [
        pasta_rreo_correspondente(state),
        uf,
        state,
    ]

    candidates: list[dict[str, Any]] = []
    used_keys: set[str] = set()

    for key in search_keys:
        key = str(key or "").strip()
        if not key or key in used_keys:
            continue

        used_keys.add(key)
        candidates = list_results(key)
        if candidates:
            break

    if candidates:
        selected = candidates[0]
        download_file(selected["blob_name"], destination)
        return destination, f"Resultado RREO mais recente: {selected['name']}"

    if not PLANILHA_BASE.exists():
        raise FileNotFoundError(
            "Nenhuma planilha RREO processada foi localizada e a "
            "planilha-base local nao foi encontrada em data/."
        )

    destination.write_bytes(PLANILHA_BASE.read_bytes())
    return destination, "Planilha-base local"


def upload_resultado_fnde(
    local_path: Path,
    state: str,
) -> dict[str, Any]:
    blob_name = (
        f"{FNDE_RESULTADOS_PREFIX}"
        f"{state}/"
        f"{local_path.name}"
    )

    return upload_file(
        local_path=local_path,
        blob_name=blob_name,
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )


@st.cache_data(ttl=300, show_spinner=False)
def carregar_estados_fnde() -> list[str]:
    return listar_estados_fnde()


@st.cache_data(ttl=180, show_spinner=False)
def carregar_pdfs_fnde(state: str) -> list[dict[str, Any]]:
    return listar_pdfs_fnde(state)


# ============================================================
# INTERFACE
# ============================================================

st.set_page_config(
    page_title="Painel FNDE",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded",
)
apply_theme()
render_sidebar()

status = health_check()
cloud_ok = bool(status.get("ok"))
gemini_ok = bool(
    (
        os.getenv("GEMINI_API_KEY")
        or os.getenv("GOOGLE_API_KEY")
        or ""
    ).strip()
)

st.markdown(
    '<div class="hero-row"><div>'
    '<div class="hero-title">Painel de Extracao FNDE</div>'
    '<div class="hero-sub">PNAE, PNATE, PDDE e QSE integrados a mesma planilha do RREO.</div>'
    '</div><span class="online">● Sistema Online</span></div>',
    unsafe_allow_html=True,
)

m1, m2, m3, m4 = st.columns(4)
with m1:
    metric_card(
        "☁", "Armazenamento", "Cloud Storage",
        "Conectado" if cloud_ok else "Verificar", "blue"
    )
with m2:
    metric_card("📄", "Fonte", "PDFs FNDE", "SIGEFWEB", "purple")
with m3:
    metric_card(
        "✦", "Gemini API",
        "Configurado" if gemini_ok else "Pendente",
        "OK" if gemini_ok else "Atencao", "green"
    )
with m4:
    metric_card("📊", "Destino", "Colunas T a W", "Excel integrado", "blue")

if not cloud_ok:
    st.error("Nao foi possivel conectar ao Google Cloud Storage.")
    st.code(status.get("message", "Erro desconhecido."))
    st.stop()

if not gemini_ok:
    st.error("Configure GEMINI_API_KEY ou GOOGLE_API_KEY no Render.")
    st.stop()

try:
    states = carregar_estados_fnde()
except Exception as error:
    st.error("Nao foi possivel listar os estados do FNDE no Cloud Storage.")
    st.exception(error)
    st.stop()

if not states:
    st.warning(
        "Nenhuma pasta FNDE foi encontrada. Use o padrao "
        "ARQUIVO_DE_ESTADOS_RREO/<Estado> FNDE - <UF>_<ANO>/."
    )
    st.stop()

st.markdown(
    '<div class="section-card"><div class="section-title">'
    '<span class="section-num">1.</span>Selecao do processamento</div>',
    unsafe_allow_html=True,
)

all_states = st.checkbox(
    "Todos os Estados FNDE",
    value=False,
    disabled=not PERMITIR_TODOS_ESTADOS,
    help="Ative esta opção em Configurações para processar todas as pastas estaduais FNDE.",
)
state = st.selectbox("Pasta estadual FNDE", states, disabled=all_states)
uf_folder, year_folder = dados_pasta_fnde(state)

c1, c2, c3 = st.columns([1.4, 1.4, 1.0])
with c1:
    st.text_input("UF identificada", value=uf_folder or "-", disabled=True)
with c2:
    mode = st.segmented_control(
        "Modo",
        options=["Estado inteiro", "Municipio unico"],
        default="Estado inteiro",
    )
with c3:
    year = st.number_input(
        "Ano identificado",
        min_value=2000,
        max_value=2100,
        value=year_folder or 2025,
        step=1,
        disabled=True,
    )

st.markdown('</div>', unsafe_allow_html=True)

if all_states:
    files = []
    for folder in states:
        for item in carregar_pdfs_fnde(folder):
            item = dict(item)
            item["state_folder"] = folder
            files.append(item)
else:
    files = carregar_pdfs_fnde(state)
    for item in files:
        item["state_folder"] = state

if not files:
    st.warning("Nenhum PDF FNDE foi encontrado nas pastas selecionadas.")
    st.stop()

selected_files = files
if all_states:
    mode = "Estado inteiro"
elif mode == "Municipio unico":
    selected_name = st.selectbox(
        "PDF do municipio",
        [item["name"] for item in files],
    )
    selected_files = [
        next(item for item in files if item["name"] == selected_name)
    ]

size_mb = sum(int(item.get("size") or 0) for item in selected_files) / (1024 * 1024)

left, middle, right = st.columns([1.45, 1.0, 1.0])

with left:
    st.markdown(
        '<div class="section-card"><div class="section-title">'
        '<span class="section-num">2.</span>Arquivos encontrados</div>',
        unsafe_allow_html=True,
    )
    a, b, c = st.columns(3)
    a.metric("PDFs", len(selected_files))
    b.metric("Tamanho", f"{size_mb:.1f} MB")
    c.metric("Ano", year)

    preview = pd.DataFrame(
        [
            {
                "Arquivo": item["name"],
                "Tamanho (KB)": round((item.get("size") or 0) / 1024, 1),
            }
            for item in selected_files[:100]
        ]
    )
    st.dataframe(
        preview,
        use_container_width=True,
        hide_index=True,
        height=330,
    )
    st.caption(
        f"Exibindo {min(100, len(selected_files))} de "
        f"{len(selected_files)} arquivo(s)."
    )
    st.markdown('</div>', unsafe_allow_html=True)

with middle:
    st.markdown(
        '<div class="section-card"><div class="section-title">'
        '<span class="section-num">3.</span>Processamento</div>',
        unsafe_allow_html=True,
    )

    job = st.session_state.setdefault(
        "fnde_job",
        {
            "status": "Aguardando",
            "progress": 0.0,
            "current": "Nenhum",
            "success": 0,
            "errors": 0,
            "total": 0,
        },
    )

    st.info(job["status"])
    st.progress(
        float(job["progress"]),
        text=f"Progresso geral: {job['progress'] * 100:.0f}%",
    )
    st.write("**Arquivo atual**")
    st.write(job["current"])

    x, y, z = st.columns(3)
    x.metric("Concluidos", job["success"])
    y.metric(
        "Pendentes",
        max(job["total"] - job["success"] - job["errors"], 0),
    )
    z.metric("Erros", job["errors"])

    start = st.button(
        "▶ Processar FNDE agora",
        type="primary",
        use_container_width=True,
    )
    st.markdown('</div>', unsafe_allow_html=True)

with right:
    st.markdown(
        '<div class="section-card"><div class="section-title">'
        '<span class="section-num">4.</span>Resultado</div>',
        unsafe_allow_html=True,
    )
    result = st.session_state.get("fnde_last_result")
    last_errors = st.session_state.get("fnde_last_errors", [])

    if result:
        st.success(result["name"])
        st.download_button(
            "⬇ Baixar planilha RREO + FNDE",
            data=result["bytes"],
            file_name=result["name"],
            mime=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            use_container_width=True,
        )
        st.caption(result["cloud"])
    elif last_errors:
        st.error("O último processamento não gerou resultado.")
        st.dataframe(
            pd.DataFrame(last_errors),
            use_container_width=True,
            hide_index=True,
        )
    else:
        st.caption("O arquivo Excel aparecera aqui apos o processamento.")
    st.markdown('</div>', unsafe_allow_html=True)


# ============================================================
# EXECUCAO
# ============================================================

if start:
    st.session_state.pop("fnde_last_result", None)
    st.session_state["fnde_last_errors"] = []

    progress = st.progress(0, text="Preparando...")
    log_box = st.empty()
    logs: list[str] = []
    errors: list[dict[str, str]] = []
    warnings_count = 0
    processed_ibge: set[str] = set()
    metrics = {
        "Data/hora de início": timestamp(),
        "Módulo": "FNDE VISUAL",
        "PDFs encontrados": len(selected_files),
        "PDFs processados": 0,
        "Municípios preenchidos": 0,
        "Municípios pendentes": 0,
        "Campos preenchidos": 0,
        "Campos vazios": 0,
        "Erros críticos": 0,
        "Avisos": 0,
        "Gemini Vision OK": 0,
        "OCR local fallback OK": 0,
        "Upload Drive": "PENDENTE",
    }

    uf, _ = dados_pasta_fnde(state)
    uf = "BRASIL" if all_states else (uf or extrair_uf(state) or state)

    with tempfile.TemporaryDirectory(prefix="fnde_cloud_") as temp_dir:
        temp = Path(temp_dir)
        source_xlsx = temp / "origem.xlsx"
        output_name = (
            f"RREO_FNDE_{uf}_"
            f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        output_xlsx = temp / output_name

        try:
            if all_states:
                source_xlsx.write_bytes(PLANILHA_BASE.read_bytes())
                source_description = "Planilha-base local (Todos os Estados FNDE)"
            else:
                _, source_description = baixar_planilha_origem(
                    state,
                    uf,
                    source_xlsx,
                )
            output_xlsx.write_bytes(source_xlsx.read_bytes())
            logs.append(f"Planilha de origem: {source_description}")

            workbook = load_workbook(output_xlsx)
            worksheet = localizar_aba_principal(workbook)
            ibge_rows = indice_linhas_ibge(worksheet)

            total = len(selected_files)
            job.update(
                {
                    "status": "Em andamento",
                    "progress": 0.0,
                    "current": "Preparando...",
                    "success": 0,
                    "errors": 0,
                    "total": total,
                }
            )

            try:
                for index, item in enumerate(selected_files, start=1):
                    job["current"] = item["name"]
                    job["progress"] = (index - 1) / total if total else 1.0

                    progress.progress(
                        job["progress"],
                        text=(
                            f"Processando {index} de {total}: "
                            f"{item['name']}"
                        ),
                    )

                    pdf_path = temp / item["name"]

                    try:
                        download_pdf(item["blob_name"], pdf_path)
                        values, _, extraction = process(
                            pdf_path,
                            enviar_imagens=True,
                        )

                        code = normalizar_codigo_ibge(extraction.codigo_ibge)
                        row = ibge_rows.get(code)

                        if row is None:
                            raise RuntimeError(
                                f"Codigo IBGE {code or 'nao identificado'} "
                                "nao encontrado na planilha."
                            )

                        preencher_linha(worksheet, row, values)
                        job["success"] += 1
                        warnings_count += len(extraction.avisos)
                        processed_ibge.add(code)
                        metrics["PDFs processados"] += 1
                        metrics["Municípios preenchidos"] += 1
                        found_programs = [p for p, v in values.items() if float(v or 0.0) != 0.0]
                        missing_programs = [p for p in COLUNAS_FNDE if p not in found_programs]
                        metrics["Campos preenchidos"] += len(found_programs)
                        metrics["Campos vazios"] += len(missing_programs)
                        metrics["Avisos"] += len(extraction.avisos)
                        if extraction.metodo == "GEMINI_VISION_OK":
                            metrics["Gemini Vision OK"] += 1
                        elif extraction.metodo == "OCR_LOCAL_FALLBACK_OK":
                            metrics["OCR local fallback OK"] += 1

                        if GERAR_LOG_FNDE:
                            append_fnde_log(workbook, {
                                "Data/Hora": timestamp(),
                                "PDF FNDE": item["name"],
                                "Município": extraction.municipio,
                                "Código IBGE": code,
                                "UF": extraction.uf or uf,
                                "Ano": year,
                                "Estado/Lote": state,
                                "Linha da planilha": row,
                                "Campos FNDE preenchidos": len(found_programs),
                                "Programas encontrados": ", ".join(found_programs),
                                "Programas ausentes": ", ".join(missing_programs),
                                "Valores extraídos": "; ".join(f"{p}={values[p]:.2f}" for p in COLUNAS_FNDE),
                                "Método de extração": extraction.metodo,
                                "Modelo Gemini": extraction.modelo,
                                "Tentativas Gemini": extraction.tentativas,
                                "Status": "OK" if found_programs else "PARCIAL",
                                "Avisos": "; ".join(extraction.avisos),
                            })

                        logs.append(
                            f"OK {code} {extraction.municipio}/{extraction.uf} "
                            f"[{extraction.metodo}]: PNAE={values['PNAE']:.2f}; "
                            f"PNATE={values['PNATE']:.2f}; PDDE={values['PDDE']:.2f}; "
                            f"QSE={values['QSE']:.2f}"
                        )

                    except Exception as error:
                        job["errors"] += 1
                        error_item = {
                            "arquivo": item["name"],
                            "erro": f"{type(error).__name__}: {error}",
                        }
                        errors.append(error_item)
                        st.session_state["fnde_last_errors"] = list(errors)

                        error_line = (
                            f"ERRO FNDE {item['name']}: "
                            f"{type(error).__name__}: {error}"
                        )
                        logs.append(error_line)
                        print(error_line, flush=True)
                        metrics["PDFs processados"] += 1
                        metrics["Erros críticos"] += 1
                        if GERAR_LOG_FNDE:
                            append_fnde_log(workbook, {
                                "Data/Hora": timestamp(), "PDF FNDE": item["name"],
                                "UF": uf, "Ano": year, "Estado/Lote": state,
                                "Status": "ERRO", "Erro resumido": error_item["erro"],
                            })

                    finally:
                        pdf_path.unlink(missing_ok=True)

                    job["progress"] = index / total if total else 1.0
                    if SALVAR_INCREMENTAL:
                        workbook.save(output_xlsx)
                        if index % CHECKPOINT_INTERVALO == 0:
                            checkpoint_blob = (
                                f"{FNDE_RESULTADOS_PREFIX}_CHECKPOINTS/"
                                f"{'BRASIL' if all_states else state}/checkpoint_{output_name}"
                            )
                            try:
                                upload_file(output_xlsx, checkpoint_blob)
                            except Exception as checkpoint_error:
                                logs.append(f"AVISO checkpoint: {checkpoint_error}")
                    log_box.code("\n".join(logs[-18:]), language=None)

                if job["success"] == 0:
                    job["status"] = "Falha"
                    job["current"] = "Nenhum PDF concluído"
                    job["progress"] = 1.0
                    st.session_state["fnde_last_errors"] = list(errors)

                    raise RuntimeError(
                        "Nenhum PDF FNDE foi processado com sucesso. "
                        "A planilha vazia não será gerada."
                    )

                if GERAR_NAO_ENCONTRADOS:
                    prefix = "" if all_states else CODIGOS_UF.get(uf, "")
                    missing_rows = []
                    code_column = localizar_coluna_ibge(worksheet)
                    for code, row_number in ibge_rows.items():
                        if prefix and not code.startswith(prefix):
                            continue
                        if code in processed_ibge:
                            continue
                        missing_rows.append({
                            "Estado/UF": uf,
                            "Código IBGE": code,
                            "Município da planilha-base": worksheet.cell(row_number, code_column + 1).value or "",
                            "Situação": "PDF FNDE NÃO PROCESSADO",
                            "PDF FNDE correspondente": "",
                            "Observação": "Não houve preenchimento FNDE seguro nesta execução.",
                        })
                    write_missing(workbook, missing_rows)
                    metrics["Municípios pendentes"] = len(missing_rows)

                metrics["Data/hora de fim"] = timestamp()
                metrics["Arquivo de saída"] = output_name
                if GERAR_AUDITORIA:
                    write_audit(workbook, metrics)
                workbook.save(output_xlsx)

            finally:
                workbook.close()

            uploaded = upload_resultado_fnde(output_xlsx, "BRASIL" if all_states else state)
            metrics["Upload Drive"] = "OK"
            result_bytes = output_xlsx.read_bytes()

            st.session_state["fnde_last_result"] = {
                "name": output_name,
                "bytes": result_bytes,
                "cloud": uploaded["blob_name"],
            }

            job["status"] = "Concluido"
            job["current"] = "Finalizado"
            job["progress"] = 1.0
            progress.progress(1.0, text="Processamento concluido.")

            st.success(
                f"Concluido: {job['success']} PDF(s), "
                f"{job['errors']} erro(s) e "
                f"{warnings_count} aviso(s)."
            )
            st.caption(f"Arquivo salvo no Cloud: {uploaded['blob_name']}")

            if errors:
                st.warning("Alguns arquivos precisam de conferencia.")
                st.dataframe(
                    pd.DataFrame(errors),
                    use_container_width=True,
                    hide_index=True,
                )

            st.rerun()

        except Exception as error:
            job["status"] = "Falha"
            job["current"] = "Interrompido"
            job["progress"] = 1.0
            st.session_state["fnde_last_errors"] = list(errors) or [
                {
                    "arquivo": "PROCESSAMENTO GERAL",
                    "erro": f"{type(error).__name__}: {error}",
                }
            ]

            general_line = (
                f"ERRO GERAL FNDE: {type(error).__name__}: {error}"
            )
            logs.append(general_line)
            print(general_line, flush=True)

            progress.progress(1.0, text="Processamento interrompido.")
            log_box.code("\n".join(logs[-18:]), language=None)
            st.error("O processamento FNDE foi interrompido.")
            st.exception(error)

            if errors:
                st.dataframe(
                    pd.DataFrame(errors),
                    use_container_width=True,
                    hide_index=True,
                )

st.markdown(
    '<div class="section-card"><div class="section-title">Fluxo Operacional</div>'
    '<div class="flow">'
    '<div class="flow-step"><div class="flow-num">1</div><div>'
    '<div class="flow-name">Listagem</div>'
    '<div class="flow-desc">PDFs FNDE localizados no Cloud Storage</div>'
    '</div></div><div class="flow-arrow">→</div>'
    '<div class="flow-step"><div class="flow-num">2</div><div>'
    '<div class="flow-name">Extracao</div>'
    '<div class="flow-desc">Gemini identifica PNAE, PNATE, PDDE e QSE</div>'
    '</div></div><div class="flow-arrow">→</div>'
    '<div class="flow-step"><div class="flow-num">3</div><div>'
    '<div class="flow-name">Preenchimento</div>'
    '<div class="flow-desc">Colunas T, U, V e W preenchidas automaticamente</div>'
    '</div></div><div class="flow-arrow">→</div>'
    '<div class="flow-step"><div class="flow-num">4</div><div>'
    '<div class="flow-name">Upload</div>'
    '<div class="flow-desc">Resultado salvo e liberado para download</div>'
    '</div></div></div></div>',
    unsafe_allow_html=True,
)
